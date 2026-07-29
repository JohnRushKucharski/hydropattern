# ruff: noqa
"""Generate examples/great_lakes/templates/template_twl.xlsx: the blank resources+config
workbook consumed by batch_run_twl.py.

One-off generator script (not part of the hydropattern package, so excluded from
linting -- see the `# ruff: noqa` above). Re-run this script any time the recognized
resources-sheet columns or config-sheet options in batch_run_twl.py change, so the
shipped template stays in sync.

Usage:
    uv run python examples/great_lakes/templates/build_twl_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

TEMPLATE_PATH = Path(__file__).parent / "template_twl.xlsx"

# Column order: required columns first, then save-point selector (id OR lat+lon),
# then optional magnitude/threshold extras.
RESOURCES_HEADER = [
    "resource_name",
    "component_name",
    "lake",
    "save_point_id",
    "lat",
    "lon",
    "success_pattern",
    "magnitude_operator",
    "magnitude_value",
    "threshold",
]

# Example rows, illustrating: save_point_id selection (row 1), lat/lon nearest-point
# selection (row 2), and a success_pattern=True low-water row (row 3). component_name
# is left blank on rows 1-2 to also illustrate its "twl" default.
EXAMPLE_RESOURCE_ROWS = [
    {
        "resource_name": "duluth_harbor",
        "lake": "superior",
        "save_point_id": 1,
        "success_pattern": False,
        "magnitude_operator": ">=",
        "magnitude_value": 183.5,
        "threshold": 10,
    },
    {
        "resource_name": "mackinac_strait",
        "component_name": "high_water",
        "lake": "huron",
        "lat": 46.5,
        "lon": -84.4,
        "success_pattern": False,
        "magnitude_operator": ">",
        "magnitude_value": 176.0,
    },
    {
        "resource_name": "kingston_shoal",
        "component_name": "low_water",
        "lake": "ontario",
        "save_point_id": 1,
        "success_pattern": True,
        "magnitude_operator": "<",
        "magnitude_value": 74.8,
    },
]

# (option, default value, help comment) -- mirrors BatchConfig's own field defaults in
# batch_run_twl.py. output_directory has no sensible default (required); its value here
# is a placeholder the user must replace.
CONFIG_ROWS = [
    ("output_directory", "output",
     "REQUIRED. Base directory for generated grid csv + plot png output files."),
    ("subdirectory_structure", "flat",
     'One of "flat", "resource", or "row". See CONTEXT.md for the "Subdirectory '
     'structure" definition.'),
    ("metric_mode", "return_period", 'One of "portion", "percentage", or "return_period".'),
    ("overwrite", False, "Allow overwriting a resource's grid csv/plot png if they already "
     "exist."),
    ("plot_interpolate", True, "Bilinearly interpolate the plotted response surface."),
    ("plot_color_map", "RdBu",
     'Matplotlib colormap name. Left at "RdBu", hydropattern auto-reverses it per '
     "component based on metric_mode/success_pattern -- see docs/user/reference.md."),
    ("plot_color_map_ticks", "",
     "(Optional) comma-separated explicit colorbar ticks, e.g. \"-1.0, 0.0, 1.0\"."),
    ("compute_equivalent_elevation", False,
     "If true, also write a second grid csv + plot png per row: the water level under "
     "every scenario equivalent (same ARI) to the baseline (_0_0) scenario's ARI at "
     "magnitude_value. See CONTEXT.md's \"Equivalent elevation\" definition."),
    ("fillin", False,
     "If true, estimate missing (NaN) grid cells in the response surface via Delaunay "
     "triangulation before plotting (climate-canvas's --fillin option). Applies to "
     "every plot png this row produces."),
]

_HEADER_FONT = Font(bold=True)

# threshold column note: unlike magnitude_value (a water-level threshold, same units as
# the twl data), 'threshold' is a climate-canvas plot reference line, expressed in the
# *metric_mode*'s output units (e.g. a portion in [0,1], a percentage in [0,100], or an
# ARI in years for return_period) -- not a water level. See docs/user/reference.md's
# `[output.plot.climate-canvas].threshold`.
_THRESHOLD_HEADER_COMMENT = Comment(
    "Climate-canvas plot reference line, in the config sheet's metric_mode units "
    "(portion 0-1, percentage 0-100, or return_period in years) -- NOT a water level.",
    "hydropattern",
)


def build_template(path: Path = TEMPLATE_PATH) -> Path:
    """Write the resources+config template workbook to path (overwriting if present)."""
    wb = Workbook()

    resources_ws = wb.active
    # A freshly created Workbook() always has one active Worksheet (never a Chartsheet
    # or None) -- openpyxl's stubs just can't express that statically.
    assert isinstance(resources_ws, Worksheet)
    resources_ws.title = "resources"
    resources_ws.append(RESOURCES_HEADER)
    for cell in resources_ws[1]:
        cell.font = _HEADER_FONT
    resources_ws.cell(row=1, column=RESOURCES_HEADER.index("threshold") + 1).comment = \
        _THRESHOLD_HEADER_COMMENT
    for example_row in EXAMPLE_RESOURCE_ROWS:
        resources_ws.append([example_row.get(col) for col in RESOURCES_HEADER])

    config_ws = wb.create_sheet("config")
    config_ws.append(["option", "value"])
    for cell in config_ws[1]:
        cell.font = _HEADER_FONT
    for option, default, help_text in CONFIG_ROWS:
        config_ws.append([option, default])
        config_ws.cell(row=config_ws.max_row, column=1).comment = Comment(help_text, "hydropattern")

    wb.save(path)
    return path


if __name__ == "__main__":
    out_path = build_template()
    print(f"Wrote template to: {out_path}")
