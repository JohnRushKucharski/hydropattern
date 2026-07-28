# ruff: noqa
"""Generate examples/great_lakes/templates/template_avg.xlsx: the blank resources+config
workbook consumed by batch_run_avg.py.

One-off generator script (not part of the hydropattern package, so excluded from
linting -- see the `# ruff: noqa` above). Re-run this script any time the recognized
resources-sheet columns or config-sheet options in batch_run_avg.py change, so the
shipped template stays in sync.

Usage:
    uv run python examples/great_lakes/templates/build_avg_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

TEMPLATE_PATH = Path(__file__).parent / "template_avg.xlsx"

# Column order matches the fixed characteristic emission order in batch_run_avg.py's
# build_toml_text (timing, magnitude, rate_of_change, duration), required columns first.
RESOURCES_HEADER = [
    "resource_name",
    "component_name",
    "lake",
    "success_pattern",
    "verbose",
    "timing_first_month",
    "timing_last_month",
    "magnitude_operator",
    "magnitude_value",
    "magnitude_ma_periods",
    "rate_of_change_operator",
    "rate_of_change_value",
    "rate_of_change_ma_periods",
    "rate_of_change_look_back",
    "rate_of_change_min_val",
    "duration_operator",
    "duration_value",
    "threshold",
]

# One filled-in example row, illustrating a magnitude-only component. Blank cells for
# columns not used by this example (timing/rate_of_change/duration/threshold).
EXAMPLE_RESOURCE_ROW = {
    "resource_name": "duluth_harbor",
    "component_name": "high_water",
    "lake": "superior",
    "success_pattern": False,
    "verbose": True,
    "magnitude_operator": ">",
    "magnitude_value": 183.0,
    "magnitude_ma_periods": 1,
}

# (option, default value, help comment) -- mirrors BatchConfig's own field defaults in
# batch_run_avg.py. output_directory has no sensible default (required); its value here
# is a placeholder the user must replace.
CONFIG_ROWS = [
    ("output_directory", "output",
     "REQUIRED. Base directory for generated .toml files and hydropattern run outputs."),
    ("subdirectory_structure", "flat",
     'One of "flat", "resource", or "row". See CONTEXT.md for the "Subdirectory '
     'structure" definition.'),
    ("first_day_of_water_year", 1, "Day-of-year (1-365) the water year starts on."),
    ("metric_mode", "portion", 'One of "portion", "percentage", or "return_period".'),
    ("excel", True, "Also write an Excel copy of each run's output."),
    ("overwrite", False, "Allow overwriting a run's output directory if it already exists."),
    ("plot_enabled", True, "Generate a climate-canvas response-surface plot per component."),
    ("plot_interpolate", True, "Bilinearly interpolate the plotted response surface."),
    ("plot_color_map", "RdBu",
     'Matplotlib colormap name. Left at "RdBu", hydropattern auto-reverses it per '
     "component based on metric_mode/success_pattern -- see docs/user/reference.md."),
    ("plot_color_map_ticks", "",
     "(Optional) comma-separated explicit colorbar ticks, e.g. \"-1.0, 0.0, 1.0\"."),
]

_HEADER_FONT = Font(bold=True)


def build_template(path: Path = TEMPLATE_PATH) -> Path:
    """Write the resources+config template workbook to path (overwriting if present)."""
    wb = Workbook()

    resources_ws = wb.active
    # A freshly created Workbook() always has one active Worksheet (never a Chartsheet
    # or None) -- openpyxl's stubs just can't express that statically.
    assert isinstance(resources_ws, Worksheet)
    resources_ws.title = "resources"
    resources_ws.append(RESOURCES_HEADER)
    for cell in resources_ws[1]:
        cell.font = _HEADER_FONT
    resources_ws.append([EXAMPLE_RESOURCE_ROW.get(col) for col in RESOURCES_HEADER])

    config_ws = wb.create_sheet("config")
    config_ws.append(["option", "value"])
    for cell in config_ws[1]:
        cell.font = _HEADER_FONT
    for option, default, help_text in CONFIG_ROWS:
        config_ws.append([option, default])
        config_ws.cell(row=config_ws.max_row, column=1).comment = Comment(help_text, "hydropattern")

    wb.save(path)
    return path


if __name__ == "__main__":
    out_path = build_template()
    print(f"Wrote template to: {out_path}")
