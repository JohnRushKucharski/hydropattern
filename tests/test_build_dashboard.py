"""Tests for the Great Lakes longtailpoint results dashboard generator.

build_dashboard.py reads one or more batch_run_twl.py resources+config workbooks and
their already-generated outputs, and produces a self-contained dashboard.html +
sidecar manifest.js (see examples/great_lakes/docs/adr/0002-sidecar-js-manifest-for-
longtailpoint-dashboard.md for why the manifest is a .js sidecar rather than a live
JSON/CSV fetch).
"""
from pathlib import Path

import pytest

from examples.great_lakes import build_dashboard as dash
from examples.great_lakes import common_twl

# ---- compute_magnitude_ft -----------------------------------------------------------

def _resource(**overrides):
    from examples.great_lakes import batch_run_twl as twl
    fields = dict(
        resource_name="longtail_17877",
        component_name="base",
        lake="michigan",
        magnitude_operator=">=",
        magnitude_value=178.7747637915269,
        save_point_id=1968,
        equivalent_elevation=None,
    )
    fields.update(overrides)
    return twl.ResourceSpec(**fields)


def test_compute_magnitude_ft_converts_m_igld85_to_ft_NAVD88():
    resource = _resource(magnitude_value=178.7747637915269)
    expected = common_twl.m_igld85_to_ft_NAVD88(178.7747637915269)
    assert dash.compute_magnitude_ft(resource) == pytest.approx(expected)


# ---- compute_equivalent_elevation_basis ----------------------------------------------

def test_basis_is_none_when_equivalent_elevation_blank():
    resource = _resource(equivalent_elevation=None)
    row = {"equivalent_elevation": None}
    assert dash.compute_equivalent_elevation_basis(row, resource) is None


def test_basis_is_baseline_magnitude_when_cell_says_so():
    resource = _resource(magnitude_value=178.77, equivalent_elevation=178.77)
    row = {"equivalent_elevation": "baseline_magnitude"}
    assert dash.compute_equivalent_elevation_basis(row, resource) == "baseline_magnitude"


def test_basis_is_baseline_magnitude_case_insensitive():
    resource = _resource(magnitude_value=178.77, equivalent_elevation=178.77)
    row = {"equivalent_elevation": "Baseline_Magnitude"}
    assert dash.compute_equivalent_elevation_basis(row, resource) == "baseline_magnitude"


def test_basis_is_override_label_in_ft_NAVD88_for_numeric_cell():
    resource = _resource(magnitude_value=178.77, equivalent_elevation=178.7)
    row = {"equivalent_elevation": 178.7}
    expected_ft = common_twl.m_igld85_to_ft_NAVD88(178.7)
    assert dash.compute_equivalent_elevation_basis(row, resource) == f"{expected_ft:.2f} ft override"


# ---- ManifestEntry / build_entries ---------------------------------------------------

def _write_workbook(
    tmp_path, output_directory, resources_rows, resources_header=None, filename_style=None,
):
    from openpyxl import Workbook

    header = resources_header or [
        "resource_name", "component_name", "lake", "magnitude_operator",
        "magnitude_value", "save_point_id", "equivalent_elevation",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "resources"
    ws.append(header)
    for row in resources_rows:
        ws.append(row)
    cfg = wb.create_sheet("config")
    cfg.append(["option", "value"])
    cfg.append(["output_directory", str(output_directory)])
    cfg.append(["metric_mode", "return_period"])
    if filename_style is not None:
        cfg.append(["filename_style", filename_style])
    path = tmp_path / "workbook.xlsx"
    wb.save(path)
    return path


def test_build_entries_reads_rows_and_resolves_output_dir(tmp_path):
    out_dir = tmp_path / "out"
    wb_path = _write_workbook(
        tmp_path, out_dir,
        [["longtail_17877", "base", "michigan", ">=", 178.7747637915269, 1968, "baseline_magnitude"]],
    )
    entries = dash.build_entries(wb_path)
    assert len(entries) == 1
    entry = entries[0]
    assert entry.resource_name == "longtail_17877"
    assert entry.component_name == "base"
    assert entry.save_point_id == 1968
    assert entry.qualified_name == "longtail_17877_base_1968"
    assert entry.output_dir == out_dir
    assert entry.equivalent_elevation_basis == "baseline_magnitude"
    assert entry.magnitude_ft == pytest.approx(common_twl.m_igld85_to_ft_NAVD88(178.7747637915269))
    assert entry.workbook_path == wb_path
    assert entry.analysis_type == "twl"


def test_build_entries_basis_none_when_blank(tmp_path):
    out_dir = tmp_path / "out"
    wb_path = _write_workbook(
        tmp_path, out_dir,
        [["longtail_1", "base", "michigan", ">=", 178.0, 1, None]],
    )
    entries = dash.build_entries(wb_path)
    assert entries[0].equivalent_elevation_basis is None


def test_build_entries_file_stem_defaults_to_qualified_name(tmp_path):
    out_dir = tmp_path / "out"
    wb_path = _write_workbook(
        tmp_path, out_dir,
        [["longtail_17877", "base", "michigan", ">=", 178.7747637915269, 1968, None]],
    )
    entries = dash.build_entries(wb_path)
    assert entries[0].file_stem == entries[0].qualified_name == "longtail_17877_base_1968"


def test_build_entries_file_stem_elevation_runup_savepoint_style(tmp_path):
    out_dir = tmp_path / "out"
    wb_path = _write_workbook(
        tmp_path, out_dir,
        [["longtail_17877", "base", "michigan", ">=", 178.7747637915269, 1968, None]],
        filename_style="elevation_runup_savepoint",
    )
    entries = dash.build_entries(wb_path)
    expected_stem = common_twl.output_file_stem(
        common_twl.m_igld85_to_ft_NAVD88(178.7747637915269), "base", 1968,
    )
    assert entries[0].file_stem == expected_stem
    assert entries[0].qualified_name == "longtail_17877_base_1968"  # identity unchanged


def test_build_entries_elevation_ft_is_magnitude_plus_runup(tmp_path):
    out_dir = tmp_path / "out"
    wb_path = _write_workbook(
        tmp_path, out_dir,
        [["longtail_17877", "run25", "michigan", ">=", 178.7747637915269, 1968, None]],
    )
    entries = dash.build_entries(wb_path)
    magnitude_ft = common_twl.m_igld85_to_ft_NAVD88(178.7747637915269)
    assert entries[0].elevation_ft == pytest.approx(magnitude_ft + 2.5)


def test_build_avg_entries_elevation_ft_equals_magnitude_ft():
    # avg entries have no runup allowance concept -- elevation_ft mirrors magnitude_ft.
    entry = dash.ManifestEntry(
        workbook_path=Path("x.xlsx"), analysis_type="avg", resource_name="r",
        component_name="c", save_point_id=None, magnitude_ft=580.0,
        equivalent_elevation_basis=None, qualified_name="r_c",
        output_dir=Path("out"), file_stem="r_c", elevation_ft=580.0,
    )
    assert entry.elevation_ft == entry.magnitude_ft


# ---- resolve_files --------------------------------------------------------------------

def _make_entry(output_dir, qualified_name="longtail_17877_base_1968", basis="baseline_magnitude", **overrides):
    fields = dict(
        workbook_path=Path("workbook.xlsx"),
        analysis_type="twl",
        resource_name="longtail_17877",
        component_name="base",
        save_point_id=1968,
        magnitude_ft=586.44,
        equivalent_elevation_basis=basis,
        qualified_name=qualified_name,
        output_dir=output_dir,
        file_stem=qualified_name,
        elevation_ft=586.44,
    )
    fields.update(overrides)
    return dash.ManifestEntry(**fields)


def _touch_all_six(output_dir, qualified_name):
    output_dir.mkdir(parents=True, exist_ok=True)
    for suffix in dash._FILE_SUFFIXES.values():
        (output_dir / f"{qualified_name}{suffix}").write_text("x", encoding="utf-8")


def test_resolve_files_returns_all_six_when_basis_set(tmp_path):
    entry = _make_entry(tmp_path)
    _touch_all_six(tmp_path, entry.qualified_name)
    files = dash.resolve_files(entry)
    assert set(files.keys()) == set(dash._FILE_SUFFIXES.keys())
    assert all(v is not None for v in files.values())


def test_resolve_files_equivalent_and_delta_kinds_none_when_basis_blank(tmp_path):
    entry = _make_entry(tmp_path, basis=None)
    tmp_path.mkdir(parents=True, exist_ok=True)
    (tmp_path / f"{entry.qualified_name}_grid.csv").write_text("x", encoding="utf-8")
    (tmp_path / f"{entry.qualified_name}_plot.png").write_text("x", encoding="utf-8")
    files = dash.resolve_files(entry)
    assert files["grid"] is not None
    assert files["plot"] is not None
    assert files["equivalent_elevation_grid"] is None
    assert files["equivalent_elevation_plot"] is None
    assert files["elevation_delta_grid"] is None
    assert files["elevation_delta_plot"] is None


def test_resolve_files_raises_on_missing_expected_file(tmp_path):
    entry = _make_entry(tmp_path)
    tmp_path.mkdir(parents=True, exist_ok=True)
    # Only touch grid+plot -- equivalent_elevation/elevation_delta files are "expected"
    # (basis is set) but missing on disk.
    (tmp_path / f"{entry.qualified_name}_grid.csv").write_text("x", encoding="utf-8")
    (tmp_path / f"{entry.qualified_name}_plot.png").write_text("x", encoding="utf-8")
    with pytest.raises(FileNotFoundError):
        dash.resolve_files(entry)


def test_resolve_files_missing_primary_plot_is_soft_not_raised(tmp_path):
    # e.g. a batch run whose plot failed (degenerate/flat grid) -- the primary grid.csv
    # exists but plot.png doesn't. resolve_files should not crash the whole dashboard
    # build; it marks the file as missing so the dashboard can show a message instead.
    entry = _make_entry(tmp_path, basis=None)
    tmp_path.mkdir(parents=True, exist_ok=True)
    (tmp_path / f"{entry.qualified_name}_grid.csv").write_text("x", encoding="utf-8")
    files = dash.resolve_files(entry)
    assert files["grid"] is not None
    assert files["plot"] == dash.MISSING


def test_build_manifest_marks_missing_primary_file_with_message(tmp_path):
    entry = _make_entry(tmp_path, basis=None)
    tmp_path.mkdir(parents=True, exist_ok=True)
    (tmp_path / f"{entry.qualified_name}_grid.csv").write_text(
        "temp_delta\\precip_delta,0.0\n0.0,1.0\n", encoding="utf-8"
    )
    manifest = dash.build_manifest([entry], dashboard_dir=tmp_path)
    file_entry = manifest["entries"][0]["files"]["plot"]
    assert file_entry["type"] == "missing"
    assert "message" in file_entry


# ---- merge_and_validate ----------------------------------------------------------------

def test_merge_and_validate_combines_unique_entries(tmp_path):
    a = _make_entry(tmp_path, qualified_name="a", save_point_id=1)
    b = _make_entry(tmp_path, qualified_name="b", save_point_id=2)
    merged = dash.merge_and_validate([[a], [b]])
    assert merged == [a, b]


def test_merge_and_validate_raises_on_duplicate_filter_combination(tmp_path):
    a = _make_entry(tmp_path, qualified_name="a", save_point_id=1, magnitude_ft=586.44)
    b = _make_entry(tmp_path, qualified_name="b", save_point_id=1, magnitude_ft=586.44)
    with pytest.raises(ValueError, match="[Nn]on-unique"):
        dash.merge_and_validate([[a], [b]])


# ---- read_grid_csv -----------------------------------------------------------------------

def test_read_grid_csv_parses_labels_and_sparse_values(tmp_path):
    csv_path = tmp_path / "grid.csv"
    csv_path.write_text(
        "temp_delta\\precip_delta,0.0,5.0\n"
        "0.0,2.5,\n"
        "1.5,1.1,2.9\n",
        encoding="utf-8",
    )
    grid = dash.read_grid_csv(csv_path)
    assert grid["col_labels"] == ["0.0", "5.0"]
    assert grid["row_labels"] == ["0.0", "1.5"]
    assert grid["values"][0][0] == pytest.approx(2.5)
    assert grid["values"][0][1] is None
    assert grid["values"][1][1] == pytest.approx(2.9)


# ---- build_manifest -------------------------------------------------------------------

def test_build_manifest_embeds_grid_data_and_relative_image_paths(tmp_path):
    out_dir = tmp_path / "out"
    entry = _make_entry(out_dir)
    _touch_all_six(out_dir, entry.qualified_name)
    (out_dir / f"{entry.qualified_name}_grid.csv").write_text(
        "temp_delta\\precip_delta,0.0\n0.0,1.0\n", encoding="utf-8"
    )
    dashboard_dir = tmp_path
    manifest = dash.build_manifest([entry], dashboard_dir)
    assert len(manifest["entries"]) == 1
    entry_data = manifest["entries"][0]
    assert entry_data["qualified_name"] == entry.qualified_name
    assert entry_data["files"]["grid"]["type"] == "grid"
    assert entry_data["files"]["grid"]["data"]["values"][0][0] == pytest.approx(1.0)
    assert entry_data["files"]["grid"]["path"] == f"out/{entry.qualified_name}_grid.csv"
    assert entry_data["files"]["plot"]["type"] == "image"
    assert entry_data["files"]["plot"]["path"] == f"out/{entry.qualified_name}_plot.png"


# ---- write_manifest_js ------------------------------------------------------------------

def test_write_manifest_js_writes_window_manifest_assignment(tmp_path):
    manifest = {"entries": [{"a": 1}]}
    path = tmp_path / "manifest.js"
    dash.write_manifest_js(manifest, path)
    text = path.read_text(encoding="utf-8")
    assert text.startswith("window.MANIFEST = ")
    assert text.strip().endswith(";")
    import json
    payload = text.strip()[len("window.MANIFEST = "):-1]
    assert json.loads(payload) == manifest


# ---- render_dashboard_html --------------------------------------------------------------

def test_render_dashboard_html_references_manifest_js(tmp_path):
    path = tmp_path / "dashboard.html"
    dash.render_dashboard_html(path)
    text = path.read_text(encoding="utf-8")
    assert '<script src="manifest.js">' in text
    assert "id=\"save-point-select\"" in text
    assert "id=\"magnitude-select\"" in text
    assert "id=\"runup-select\"" in text
    assert "id=\"basis-select\"" in text
    assert "id=\"form-select\"" in text
    assert "id=\"metric-select\"" in text
    assert "id=\"compare-toggle\"" in text


# ---- CLI main / build_dashboard integration ----------------------------------------------

def test_build_dashboard_end_to_end_writes_manifest_and_html(tmp_path):
    out_dir = tmp_path / "results"
    wb_path = _write_workbook(
        tmp_path, out_dir,
        [["longtail_1", "base", "michigan", ">=", 178.0, 1, "baseline_magnitude"]],
    )
    _touch_all_six(out_dir, "longtail_1_base_1")
    (out_dir / "longtail_1_base_1_grid.csv").write_text(
        "temp_delta\\precip_delta,0.0\n0.0,1.0\n", encoding="utf-8"
    )
    dashboard_dir = tmp_path / "dashboard"
    dash.build_dashboard(twl_workbook_paths=[wb_path], avg_workbook_paths=[], output_dir=dashboard_dir)
    assert (dashboard_dir / "manifest.js").exists()
    assert (dashboard_dir / "dashboard.html").exists()


# ---- build_avg_entries (average lake level runs) -------------------------------------

def _write_avg_workbook(tmp_path, output_directory, resources_rows, resources_header=None):
    from openpyxl import Workbook

    header = resources_header or [
        "resource_name", "component_name", "lake",
        "magnitude_operator", "magnitude_value", "magnitude_ma_periods",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "resources"
    ws.append(header)
    for row in resources_rows:
        ws.append(row)
    cfg = wb.create_sheet("config")
    cfg.append(["option", "value"])
    cfg.append(["output_directory", str(output_directory)])
    cfg.append(["metric_mode", "portion"])
    path = tmp_path / "avg_workbook.xlsx"
    wb.save(path)
    return path


def test_build_avg_entries_reads_rows_and_marks_analysis_type_avg(tmp_path):
    out_dir = tmp_path / "avg_out"
    wb_path = _write_avg_workbook(
        tmp_path, out_dir,
        [["longtail_17877", "mo_avg", "michiganhuron", ">=", 178.7747637915269, 1]],
    )
    entries = dash.build_avg_entries(wb_path)
    assert len(entries) == 1
    entry = entries[0]
    assert entry.analysis_type == "avg"
    assert entry.resource_name == "longtail_17877"
    assert entry.component_name == "mo_avg"
    assert entry.save_point_id is None
    assert entry.equivalent_elevation_basis is None
    assert entry.qualified_name == "longtail_17877_mo_avg"
    assert entry.output_dir == out_dir
    assert entry.magnitude_ft == pytest.approx(common_twl.m_igld85_to_ft_NAVD88(178.7747637915269))


def test_build_avg_entries_requires_magnitude_characteristic(tmp_path):
    out_dir = tmp_path / "avg_out"
    wb_path = _write_avg_workbook(
        tmp_path, out_dir,
        [["longtail_1", "mo_avg", "michiganhuron", None, None, None, 1, 3]],
        resources_header=[
            "resource_name", "component_name", "lake",
            "magnitude_operator", "magnitude_value", "magnitude_ma_periods",
            "timing_first_month", "timing_last_month",
        ],
    )
    with pytest.raises(ValueError, match="magnitude"):
        dash.build_avg_entries(wb_path)


# ---- merge_and_validate distinguishes analysis_type -----------------------------------

def test_merge_and_validate_allows_same_identifiers_across_different_analysis_types(tmp_path):
    twl_entry = _make_entry(tmp_path, qualified_name="twl_a", analysis_type="twl",
                             save_point_id=None, magnitude_ft=500.0, component_name="mo_avg",
                             equivalent_elevation_basis=None)
    avg_entry = _make_entry(tmp_path, qualified_name="avg_a", analysis_type="avg",
                             save_point_id=None, magnitude_ft=500.0, component_name="mo_avg",
                             equivalent_elevation_basis=None)
    merged = dash.merge_and_validate([[twl_entry], [avg_entry]])
    assert merged == [twl_entry, avg_entry]


def test_merge_and_validate_raises_on_duplicate_within_same_analysis_type(tmp_path):
    a = _make_entry(tmp_path, qualified_name="a", analysis_type="avg", save_point_id=None,
                     magnitude_ft=500.0, component_name="mo_avg", equivalent_elevation_basis=None)
    b = _make_entry(tmp_path, qualified_name="b", analysis_type="avg", save_point_id=None,
                     magnitude_ft=500.0, component_name="mo_avg", equivalent_elevation_basis=None)
    with pytest.raises(ValueError, match="[Nn]on-unique"):
        dash.merge_and_validate([[a], [b]])


# ---- dashboard.html includes analysis-type picker --------------------------------------

def test_render_dashboard_html_includes_analysis_select(tmp_path):
    path = tmp_path / "dashboard.html"
    dash.render_dashboard_html(path)
    text = path.read_text(encoding="utf-8")
    assert 'id="analysis-select"' in text


def test_render_dashboard_html_includes_cascading_facet_narrowing(tmp_path):
    # The 4 identifying-filter dropdowns must narrow to only valid combinations as the
    # user picks values, rather than always listing every value for the analysis type
    # -- otherwise a user could pick a combination with no matching row.
    path = tmp_path / "dashboard.html"
    dash.render_dashboard_html(path)
    text = path.read_text(encoding="utf-8")
    assert "refreshFacetSelects" in text
    assert "matchesFilters" in text


def test_render_dashboard_html_facet_narrowing_does_not_require_prior_selection(tmp_path):
    # On first page load every facet <select> starts with zero <option>s (none
    # populated yet). matchesFilters/currentFieldValues must not treat that
    # not-yet-populated state as "value == unmatchable empty string" for every other
    # facet (which would make every facet compute zero allowed options on init) --
    # it must only constrain using selects that already have real options/selections.
    path = tmp_path / "dashboard.html"
    dash.render_dashboard_html(path)
    text = path.read_text(encoding="utf-8")
    assert "options.length" in text


def test_render_dashboard_html_resets_facets_on_analysis_type_switch(tmp_path):
    # Switching analysis type must clear stale facet <option>s from the other
    # analysis type before recomputing -- otherwise a leftover selected value (e.g. an
    # avg row's save_point_id) wrongly zeroes out every facet for the new type.
    path = tmp_path / "dashboard.html"
    dash.render_dashboard_html(path)
    text = path.read_text(encoding="utf-8")
    assert "resetFacetSelects" in text


def test_render_dashboard_html_includes_second_picker_set_for_compare(tmp_path):
    # Compare mode needs its own independently-cascading set of the 4 identifying
    # filters (ids suffixed "-b") so a user can actually choose what appears in panel
    # B, not just see it resized/revealed with panel A's selections.
    path = tmp_path / "dashboard.html"
    dash.render_dashboard_html(path)
    text = path.read_text(encoding="utf-8")
    for select_id in (
        "save-point-select-b", "magnitude-select-b", "runup-select-b", "basis-select-b",
    ):
        assert f'id="{select_id}"' in text


def test_render_dashboard_html_formats_grid_values_to_two_decimals(tmp_path):
    # Grid table cells previously rendered raw floats (varying decimal lengths), which
    # overlap/collide visually when two grids are shown side by side in compare mode.
    path = tmp_path / "dashboard.html"
    dash.render_dashboard_html(path)
    text = path.read_text(encoding="utf-8")
    assert "toFixed(2)" in text


def test_render_dashboard_html_shows_relative_file_path_caption(tmp_path):
    # Once a plot/grid is displayed, show the resolved file's path relative to
    # dashboard.html's own directory (not an absolute path) as a small caption.
    path = tmp_path / "dashboard.html"
    dash.render_dashboard_html(path)
    text = path.read_text(encoding="utf-8")
    assert "file-path-caption" in text


# ---- CLI: separate --twl-workbooks / --avg-workbooks -----------------------------------

def test_main_cli_accepts_twl_and_avg_workbooks(tmp_path):
    twl_out = tmp_path / "twl_out"
    twl_wb = _write_workbook(
        tmp_path, twl_out,
        [["longtail_1", "base", "michigan", ">=", 178.0, 1, "baseline_magnitude"]],
    )
    _touch_all_six(twl_out, "longtail_1_base_1")
    (twl_out / "longtail_1_base_1_grid.csv").write_text(
        "temp_delta\\precip_delta,0.0\n0.0,1.0\n", encoding="utf-8"
    )

    avg_out = tmp_path / "avg_out"
    avg_wb = _write_avg_workbook(
        tmp_path, avg_out,
        [["longtail_2", "mo_avg", "michiganhuron", ">=", 178.0, 1]],
    )
    (avg_out).mkdir(parents=True, exist_ok=True)
    (avg_out / "longtail_2_mo_avg_grid.csv").write_text(
        "temp_delta\\precip_delta,0.0\n0.0,1.0\n", encoding="utf-8"
    )
    (avg_out / "longtail_2_mo_avg_plot.png").write_text("x", encoding="utf-8")

    dashboard_dir = tmp_path / "dashboard"
    exit_code = dash.main([
        "--twl-workbooks", str(twl_wb),
        "--avg-workbooks", str(avg_wb),
        "--output-dir", str(dashboard_dir),
    ])
    assert exit_code == 0
    import json
    text = (dashboard_dir / "manifest.js").read_text(encoding="utf-8")
    payload = text[len("window.MANIFEST = "):].rstrip().rstrip(";")
    manifest = json.loads(payload)
    analysis_types = {e["analysis_type"] for e in manifest["entries"]}
    assert analysis_types == {"twl", "avg"}


def test_main_cli_invokes_build_dashboard(tmp_path, monkeypatch):
    out_dir = tmp_path / "results"
    wb_path = _write_workbook(
        tmp_path, out_dir,
        [["longtail_1", "base", "michigan", ">=", 178.0, 1, "baseline_magnitude"]],
    )
    _touch_all_six(out_dir, "longtail_1_base_1")
    (out_dir / "longtail_1_base_1_grid.csv").write_text(
        "temp_delta\\precip_delta,0.0\n0.0,1.0\n", encoding="utf-8"
    )
    dashboard_dir = tmp_path / "dashboard"
    exit_code = dash.main(["--twl-workbooks", str(wb_path), "--output-dir", str(dashboard_dir)])
    assert exit_code == 0
    assert (dashboard_dir / "dashboard.html").exists()
