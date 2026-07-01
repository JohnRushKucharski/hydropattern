'''Tests for formatter metric computation.'''

import unittest
from pathlib import Path

import numpy as np
import pandas as pd

from hydropattern.formatters import (
    build_summary_sheet,
    compute_metric_series,
    compute_portion_series,
    write_summary,
)
from hydropattern.parsers import MetricMode
from hydropattern.patterns import CharacteristicType, Component, Characteristic, Result


def _make_result(values: list[int], years: list[int],
                 char_name: str = 'magnitude',
                 comp_name: str = 'comp') -> Result:
    '''Build a minimal Result using Jan 1 timestamps (WY = CY).'''
    index = pd.DatetimeIndex(
        [pd.Timestamp(f'{y}-01-01') for y in years], name='time'
    )
    return _make_result_with_dates(values, index, char_name=char_name, comp_name=comp_name)


def _make_result_with_dates(values: list[int], index: pd.DatetimeIndex,
                             char_name: str = 'magnitude',
                             comp_name: str = 'comp') -> Result:
    '''Build a minimal Result with arbitrary timestamps.'''
    dowy = list(range(1, len(values) + 1))
    df = pd.DataFrame({
        'dv': [1.0] * len(values),
        'dowy': dowy,
        char_name: values,
        comp_name: values,
    }, index=index)
    char = Characteristic(
        name=char_name,
        fx=lambda df, out: np.array(values),
        type=CharacteristicType.MAGNITUDE,
    )
    component = Component(name=comp_name, characteristics=[char], is_success_pattern=True)
    result = object.__new__(Result)
    result.df = df
    result.dv_name = 'dv'
    result.component = component
    return result


class TestComputePortionSeries(unittest.TestCase):
    '''Tests for compute_portion_series().'''

    def test_all_success_total_is_one(self):
        '''All successes -> total portion = 1.0.'''
        result = _make_result([1, 1, 1, 1], [2000, 2000, 2001, 2001])
        s = compute_portion_series(result, 'magnitude')
        self.assertAlmostEqual(s['total'], 1.0)

    def test_no_success_total_is_zero(self):
        '''Zero successes -> total portion = 0.0 (not blank).'''
        result = _make_result([0, 0, 0, 0], [2000, 2000, 2001, 2001])
        s = compute_portion_series(result, 'magnitude')
        self.assertAlmostEqual(s['total'], 0.0)

    def test_partial_success_total_portion(self):
        '''2 of 4 successes -> total portion = 0.5.'''
        result = _make_result([1, 0, 1, 0], [2000, 2000, 2001, 2001])
        s = compute_portion_series(result, 'magnitude')
        self.assertAlmostEqual(s['total'], 0.5)

    def test_per_water_year_rows(self):
        '''Series index contains total + one entry per distinct water year.'''
        result = _make_result([1, 1, 0, 0], [2000, 2000, 2001, 2001])
        s = compute_portion_series(result, 'magnitude')
        self.assertIn('total', s.index)
        self.assertIn(2000, s.index)
        self.assertIn(2001, s.index)

    def test_per_water_year_values(self):
        '''Per-year portions computed correctly.'''
        result = _make_result([1, 1, 0, 0], [2000, 2000, 2001, 2001])
        s = compute_portion_series(result, 'magnitude')
        self.assertAlmostEqual(s[2000], 1.0)
        self.assertAlmostEqual(s[2001], 0.0)

    def test_zero_year_portion_is_zero_not_na(self):
        '''Year with zero successes -> 0.0, not NaN/None.'''
        result = _make_result([1, 1, 0, 0], [2000, 2000, 2001, 2001])
        s = compute_portion_series(result, 'magnitude')
        self.assertFalse(pd.isna(s[2001]))
        self.assertAlmostEqual(s[2001], 0.0)

    def test_component_column(self):
        '''Works for the component column, not just a characteristic.'''
        result = _make_result([1, 0, 1, 0], [2000, 2000, 2001, 2001], comp_name='mycomp')
        s = compute_portion_series(result, 'mycomp')
        self.assertAlmostEqual(s['total'], 0.5)

    def test_single_year(self):
        '''Single water year -> index is [total, year].'''
        result = _make_result([1, 0, 1], [2005, 2005, 2005])
        s = compute_portion_series(result, 'magnitude')
        self.assertEqual(list(s.index), ['total', 2005])
        self.assertAlmostEqual(s['total'], 2 / 3)
        self.assertAlmostEqual(s[2005], 2 / 3)


class TestComputePortionSeriesWaterYear(unittest.TestCase):
    '''Tests for compute_portion_series() with non-calendar water years.'''

    # first_day_of_wy=274 = Oct 1 (US water year).
    # WY label = ending calendar year (US convention).
    # Oct 1 1970 (doy 274) -> WY 1971; Jan 1 1971 (doy 1) -> WY 1971.

    def _oct_jan_result(self) -> Result:
        '''Two records: Oct 1 1970 (success) and Jan 1 1971 (failure).'''
        index = pd.DatetimeIndex([
            pd.Timestamp('1970-10-01'),  # doy 274 -> WY 1971
            pd.Timestamp('1971-01-01'),  # doy   1 -> WY 1971
        ], name='time')
        return _make_result_with_dates([1, 0], index)

    def test_oct_jan_grouped_into_same_wy(self):
        '''Oct 1970 and Jan 1971 both fall in WY 1971 with first_day_of_wy=274.'''
        result = self._oct_jan_result()
        s = compute_portion_series(result, 'magnitude', first_day_of_wy=274)
        self.assertIn(1971, s.index)
        self.assertNotIn(1970, s.index)

    def test_oct_jan_wy_portion(self):
        '''1 success + 1 failure in WY 1971 -> portion 0.5.'''
        result = self._oct_jan_result()
        s = compute_portion_series(result, 'magnitude', first_day_of_wy=274)
        self.assertAlmostEqual(s[1971], 0.5)

    def test_sep_and_oct_in_different_wys(self):
        '''Sep 30 1970 (doy 273) -> WY 1970; Oct 1 1970 (doy 274) -> WY 1971.'''
        index = pd.DatetimeIndex([
            pd.Timestamp('1970-09-30'),  # doy 273 < 274 -> WY 1970
            pd.Timestamp('1970-10-01'),  # doy 274 >= 274 -> WY 1971
        ], name='time')
        result = _make_result_with_dates([1, 1], index)
        s = compute_portion_series(result, 'magnitude', first_day_of_wy=274)
        self.assertIn(1970, s.index)
        self.assertIn(1971, s.index)

    def test_default_first_day_of_wy_is_calendar_year(self):
        '''Default first_day_of_wy=1 -> WY label = calendar year.'''
        result = _make_result([1, 0], [2000, 2001])
        s = compute_portion_series(result, 'magnitude')
        self.assertIn(2000, s.index)
        self.assertIn(2001, s.index)


class TestComputeMetricSeries(unittest.TestCase):
    '''Tests for compute_metric_series() — metric-mode transform on top of portion.'''

    def test_portion_mode_matches_compute_portion_series(self):
        '''PORTION mode is a passthrough of compute_portion_series.'''
        result = _make_result([1, 0, 1, 0], [2000, 2000, 2001, 2001])
        expected = compute_portion_series(result, 'magnitude')
        actual = compute_metric_series(result, 'magnitude', MetricMode.PORTION)
        pd.testing.assert_series_equal(actual, expected)

    def test_default_mode_is_portion(self):
        '''Default mode argument behaves like PORTION.'''
        result = _make_result([1, 1, 0, 0], [2000, 2000, 2001, 2001])
        s = compute_metric_series(result, 'magnitude')
        self.assertAlmostEqual(s['total'], 0.5)

    def test_percentage_mode_scales_by_100(self):
        '''PERCENTAGE mode = portion * 100.'''
        result = _make_result([1, 0, 1, 0], [2000, 2000, 2001, 2001])
        s = compute_metric_series(result, 'magnitude', MetricMode.PERCENTAGE)
        self.assertAlmostEqual(s['total'], 50.0)

    def test_percentage_mode_zero_success_is_zero(self):
        '''PERCENTAGE mode: zero successes -> 0.0, not NA.'''
        result = _make_result([0, 0, 0, 0], [2000, 2000, 2001, 2001])
        s = compute_metric_series(result, 'magnitude', MetricMode.PERCENTAGE)
        self.assertAlmostEqual(s['total'], 0.0)

    def test_return_period_mode_is_inverse_of_portion(self):
        '''RETURN_PERIOD mode = 1 / portion.'''
        result = _make_result([1, 0, 0, 0], [2000, 2000, 2001, 2001])
        s = compute_metric_series(result, 'magnitude', MetricMode.RETURN_PERIOD)
        self.assertAlmostEqual(s['total'], 4.0)

    def test_return_period_mode_zero_portion_is_na_not_inf(self):
        '''RETURN_PERIOD mode: zero-success portion -> NA, never inf (NA/zero policy).'''
        result = _make_result([0, 0, 0, 0], [2000, 2000, 2001, 2001])
        s = compute_metric_series(result, 'magnitude', MetricMode.RETURN_PERIOD)
        self.assertTrue(pd.isna(s['total']))

    def test_return_period_mode_preserves_existing_na(self):
        '''RETURN_PERIOD mode: an already-NA portion (T=0 group) stays NA.'''
        result = _make_result([1, 1], [2000, 2000])
        s = compute_metric_series(result, 'magnitude', MetricMode.RETURN_PERIOD)
        self.assertTrue(pd.isna(s.get(2001, pd.NA)))


class TestBuildSummarySheet(unittest.TestCase):
    '''Tests for build_summary_sheet().

    build_summary_sheet(scenario_results, component_name, column, first_day_of_wy=1)
    -> DataFrame: index=['total', wy...], columns=scenario_names.
    '''

    def _two_scenario_results(self) -> dict[str, list[Result]]:
        '''Two scenarios, same years, different success patterns.'''
        r_a = _make_result([1, 1, 0, 0], [2000, 2000, 2001, 2001])
        r_b = _make_result([0, 0, 1, 1], [2000, 2000, 2001, 2001])
        return {'scenario_a': [r_a], 'scenario_b': [r_b]}

    def test_columns_are_scenario_names(self):
        '''DataFrame columns = scenario names in insertion order.'''
        sr = self._two_scenario_results()
        df = build_summary_sheet(sr, 'comp', 'magnitude')
        self.assertEqual(list(df.columns), ['scenario_a', 'scenario_b'])

    def test_index_starts_with_total(self):
        '''First row index is "total".'''
        sr = self._two_scenario_results()
        df = build_summary_sheet(sr, 'comp', 'magnitude')
        self.assertEqual(df.index[0], 'total')

    def test_index_contains_water_years(self):
        '''Index contains all water years present in the data.'''
        sr = self._two_scenario_results()
        df = build_summary_sheet(sr, 'comp', 'magnitude')
        self.assertIn(2000, df.index)
        self.assertIn(2001, df.index)

    def test_total_row_values(self):
        '''Total row = portion across full dataset per scenario.'''
        sr = self._two_scenario_results()
        df = build_summary_sheet(sr, 'comp', 'magnitude')
        self.assertAlmostEqual(df.loc['total', 'scenario_a'], 0.5)
        self.assertAlmostEqual(df.loc['total', 'scenario_b'], 0.5)

    def test_per_year_values(self):
        '''Per-year values match compute_portion_series.'''
        sr = self._two_scenario_results()
        df = build_summary_sheet(sr, 'comp', 'magnitude')
        self.assertAlmostEqual(df.loc[2000, 'scenario_a'], 1.0)
        self.assertAlmostEqual(df.loc[2001, 'scenario_a'], 0.0)
        self.assertAlmostEqual(df.loc[2000, 'scenario_b'], 0.0)
        self.assertAlmostEqual(df.loc[2001, 'scenario_b'], 1.0)

    def test_single_scenario(self):
        '''Single scenario -> single column DataFrame.'''
        r = _make_result([1, 0, 1, 0], [2000, 2000, 2001, 2001])
        df = build_summary_sheet({'only': [r]}, 'comp', 'magnitude')
        self.assertEqual(list(df.columns), ['only'])
        self.assertAlmostEqual(df.loc['total', 'only'], 0.5)

    def test_component_column(self):
        '''Works for the component column, not just a characteristic.'''
        r = _make_result([1, 0, 1, 0], [2000, 2000, 2001, 2001])
        df = build_summary_sheet({'s': [r]}, 'comp', 'comp')
        self.assertAlmostEqual(df.loc['total', 's'], 0.5)

    def test_metric_mode_is_threaded_through(self):
        '''mode parameter controls the metric computed for the whole sheet.'''
        r = _make_result([1, 0, 1, 0], [2000, 2000, 2001, 2001])
        df = build_summary_sheet({'s': [r]}, 'comp', 'magnitude', mode=MetricMode.PERCENTAGE)
        self.assertAlmostEqual(df.loc['total', 's'], 50.0)

    def test_non_default_water_year(self):
        '''first_day_of_wy is passed through to compute_portion_series.'''
        index = pd.DatetimeIndex([
            pd.Timestamp('1970-10-01'),  # WY 1971 with first_day_of_wy=274
            pd.Timestamp('1971-01-01'),  # WY 1971
        ], name='time')
        r = _make_result_with_dates([1, 0], index)
        df = build_summary_sheet({'s': [r]}, 'comp', 'magnitude', first_day_of_wy=274)
        self.assertIn(1971, df.index)
        self.assertNotIn(1970, df.index)
        self.assertAlmostEqual(df.loc[1971, 's'], 0.5)


class TestWriteSummary(unittest.TestCase):
    '''Tests for write_summary() — file output of per-component summary Excel files.'''

    def _scenario_results(self) -> dict[str, list[Result]]:
        r_a = _make_result([1, 1, 0, 0], [2000, 2000, 2001, 2001])
        r_b = _make_result([0, 0, 1, 1], [2000, 2000, 2001, 2001])
        return {'scenario_a': [r_a], 'scenario_b': [r_b]}

    def test_creates_summary_file_per_component(self):
        '''One {component}_summary.xlsx file per component in output dir.'''
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp)
            write_summary(self._scenario_results(), output_path)
            self.assertTrue((output_path / 'comp_summary.xlsx').exists())

    def test_summary_has_sheet_per_characteristic_and_component(self):
        '''Summary xlsx has one sheet per characteristic + one for the component.'''
        import tempfile
        from openpyxl import load_workbook
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp)
            write_summary(self._scenario_results(), output_path)
            wb = load_workbook(output_path / 'comp_summary.xlsx', read_only=True)
            sheetnames = wb.sheetnames
            wb.close()
            # component 'comp' has one characteristic 'magnitude' + component itself
            self.assertIn('magnitude', sheetnames)
            self.assertIn('comp', sheetnames)

    def test_summary_sheet_columns_are_scenarios(self):
        '''Each summary sheet has scenario names as columns.'''
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp)
            write_summary(self._scenario_results(), output_path)
            df = pd.read_excel(output_path / 'comp_summary.xlsx',
                               sheet_name='magnitude', index_col=0)
            self.assertIn('scenario_a', df.columns)
            self.assertIn('scenario_b', df.columns)

    def test_summary_total_row_values(self):
        '''Total row in summary sheet matches expected portion.'''
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp)
            write_summary(self._scenario_results(), output_path)
            df = pd.read_excel(output_path / 'comp_summary.xlsx',
                               sheet_name='magnitude', index_col=0)
            self.assertAlmostEqual(df.loc['total', 'scenario_a'], 0.5)
            self.assertAlmostEqual(df.loc['total', 'scenario_b'], 0.5)

    def test_write_summary_overwrite_replaces_file(self):
        '''Default overwrite=True replaces existing summary file.'''
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp)
            write_summary(self._scenario_results(), output_path)
            write_summary(self._scenario_results(), output_path)
            files = list(output_path.glob('comp_summary*.xlsx'))
            self.assertEqual(len(files), 1)

    def test_write_summary_no_overwrite_appends_suffix(self):
        '''overwrite=False appends __1 suffix instead of replacing.'''
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp)
            write_summary(self._scenario_results(), output_path, overwrite=False)
            write_summary(self._scenario_results(), output_path, overwrite=False)
            self.assertTrue((output_path / 'comp_summary.xlsx').exists())
            self.assertTrue((output_path / 'comp_summary__1.xlsx').exists())
