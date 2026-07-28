"""Tests for the Great Lakes batch_run_avg.py example script.

The script lives outside the hydropattern package (examples/great_lakes/), so it is
loaded here via importlib by file path rather than a normal package import.
"""
import importlib.util
import sys
import tomllib
from dataclasses import replace
from pathlib import Path

import pytest

SCRIPT_PATH = Path(__file__).parent.parent / "examples" / "great_lakes" / "batch_run_avg.py"
_spec = importlib.util.spec_from_file_location("avg_batch_run", SCRIPT_PATH)
assert _spec is not None and _spec.loader is not None
avg_batch_run = importlib.util.module_from_spec(_spec)
sys.modules["avg_batch_run"] = avg_batch_run
_spec.loader.exec_module(avg_batch_run)


# ---- month_to_doy -----------------------------------------------------------------

@pytest.mark.parametrize("month,expected_doy", [
    (1, 1),
    (2, 32),
    (3, 60),
    (11, 305),
    (12, 335),
])
def test_month_to_doy(month, expected_doy):
    assert avg_batch_run.month_to_doy(month) == expected_doy


@pytest.mark.parametrize("month", [0, 13, -1])
def test_month_to_doy_out_of_range_raises(month):
    with pytest.raises(ValueError):
        avg_batch_run.month_to_doy(month)


# ---- parse_resource_row: required fields -------------------------------------------

def base_row(**overrides):
    row = {
        "resource_name": "duluth_harbor",
        "component_name": "high_water",
        "lake": "superior",
        "magnitude_operator": ">",
        "magnitude_value": 183.0,
    }
    row.update(overrides)
    return row


def test_parse_resource_row_minimal_valid_row():
    spec = avg_batch_run.parse_resource_row(base_row())
    assert spec.resource_name == "duluth_harbor"
    assert spec.component_name == "high_water"
    assert spec.lake == "superior"
    assert spec.success_pattern is False
    assert spec.verbose is True
    assert spec.magnitude == (">", 183.0, 1)
    assert spec.timing is None
    assert spec.rate_of_change is None
    assert spec.duration is None
    assert spec.threshold is None


@pytest.mark.parametrize("missing", ["resource_name", "component_name", "lake"])
def test_parse_resource_row_missing_required_field_raises(missing):
    row = base_row()
    row[missing] = ""
    with pytest.raises(avg_batch_run.RowValidationError):
        avg_batch_run.parse_resource_row(row)


def test_parse_resource_row_unknown_lake_raises():
    row = base_row(lake="michigan")  # not a valid lake code
    with pytest.raises(avg_batch_run.RowValidationError) as exc_info:
        avg_batch_run.parse_resource_row(row)
    assert "lake" in str(exc_info.value).lower()


def test_parse_resource_row_no_characteristics_raises():
    row = base_row()
    del row["magnitude_operator"]
    del row["magnitude_value"]
    with pytest.raises(avg_batch_run.RowValidationError) as exc_info:
        avg_batch_run.parse_resource_row(row)
    assert "characteristic" in str(exc_info.value).lower()


# ---- parse_resource_row: success_pattern / verbose ---------------------------------

@pytest.mark.parametrize("raw,expected", [(True, True), (False, False), ("", False), (None, False)])
def test_parse_resource_row_success_pattern_blank_defaults_false(raw, expected):
    row = base_row(success_pattern=raw)
    spec = avg_batch_run.parse_resource_row(row)
    assert spec.success_pattern is expected


@pytest.mark.parametrize("raw,expected", [(True, True), (False, False), ("", True), (None, True)])
def test_parse_resource_row_verbose_blank_defaults_true(raw, expected):
    row = base_row(verbose=raw)
    spec = avg_batch_run.parse_resource_row(row)
    assert spec.verbose is expected


# ---- parse_resource_row: timing (month -> doy, incl. wraparound) -------------------

def test_parse_resource_row_timing_converts_months_to_doy():
    row = base_row(timing_first_month=11, timing_last_month=12)
    spec = avg_batch_run.parse_resource_row(row)
    assert spec.timing == (305, 335)


def test_parse_resource_row_timing_wraparound_allowed():
    # Nov (11) -> Feb (2) wraps the year boundary; hydropattern core already supports
    # first_doy > last_doy, so this must NOT raise.
    row = base_row(timing_first_month=11, timing_last_month=2)
    spec = avg_batch_run.parse_resource_row(row)
    assert spec.timing == (305, 32)


def test_parse_resource_row_timing_requires_both_months():
    row = base_row(timing_first_month=11)
    with pytest.raises(avg_batch_run.RowValidationError):
        avg_batch_run.parse_resource_row(row)


@pytest.mark.parametrize("month", [0, 13])
def test_parse_resource_row_timing_month_out_of_range_raises(month):
    row = base_row(timing_first_month=month, timing_last_month=6)
    with pytest.raises(avg_batch_run.RowValidationError):
        avg_batch_run.parse_resource_row(row)


# ---- parse_resource_row: magnitude --------------------------------------------------

def test_parse_resource_row_magnitude_ma_periods_default():
    spec = avg_batch_run.parse_resource_row(base_row())
    assert spec.magnitude == (">", 183.0, 1)


def test_parse_resource_row_magnitude_ma_periods_explicit():
    row = base_row(magnitude_ma_periods=3)
    spec = avg_batch_run.parse_resource_row(row)
    assert spec.magnitude == (">", 183.0, 3)


def test_parse_resource_row_magnitude_requires_value():
    row = base_row()
    del row["magnitude_value"]
    with pytest.raises(avg_batch_run.RowValidationError):
        avg_batch_run.parse_resource_row(row)


def test_parse_resource_row_magnitude_invalid_operator_raises():
    row = base_row(magnitude_operator="~")
    with pytest.raises(avg_batch_run.RowValidationError):
        avg_batch_run.parse_resource_row(row)


# ---- parse_resource_row: rate_of_change --------------------------------------------

def test_parse_resource_row_rate_of_change_defaults():
    row = base_row(rate_of_change_operator=">", rate_of_change_value=2.0)
    del row["magnitude_operator"]
    del row["magnitude_value"]
    spec = avg_batch_run.parse_resource_row(row)
    assert spec.rate_of_change == (">", 2.0, 1, 1, 0.0)


def test_parse_resource_row_rate_of_change_explicit_params():
    row = base_row(
        rate_of_change_operator=">",
        rate_of_change_value=2.0,
        rate_of_change_ma_periods=3,
        rate_of_change_look_back=2,
        rate_of_change_min_val=0.5,
    )
    del row["magnitude_operator"]
    del row["magnitude_value"]
    spec = avg_batch_run.parse_resource_row(row)
    assert spec.rate_of_change == (">", 2.0, 3, 2, 0.5)


def test_parse_resource_row_rate_of_change_requires_value():
    row = base_row(rate_of_change_operator=">")
    with pytest.raises(avg_batch_run.RowValidationError):
        avg_batch_run.parse_resource_row(row)


# ---- parse_resource_row: duration --------------------------------------------------

def test_parse_resource_row_duration():
    row = base_row(duration_operator=">", duration_value=3)
    del row["magnitude_operator"]
    del row["magnitude_value"]
    spec = avg_batch_run.parse_resource_row(row)
    assert spec.duration == (">", 3)


def test_parse_resource_row_duration_requires_value():
    row = base_row(duration_operator=">")
    with pytest.raises(avg_batch_run.RowValidationError):
        avg_batch_run.parse_resource_row(row)


# ---- parse_resource_row: threshold --------------------------------------------------

def test_parse_resource_row_threshold_optional():
    row = base_row(threshold=183.5)
    spec = avg_batch_run.parse_resource_row(row)
    assert spec.threshold == 183.5


def test_parse_resource_row_threshold_blank_is_none():
    spec = avg_batch_run.parse_resource_row(base_row())
    assert spec.threshold is None


# ---- parse_resource_row: multiple characteristics together -------------------------

def test_parse_resource_row_multiple_characteristics():
    row = base_row(
        timing_first_month=11, timing_last_month=12,
        rate_of_change_operator=">", rate_of_change_value=2.0,
        duration_operator=">", duration_value=3,
    )
    spec = avg_batch_run.parse_resource_row(row)
    assert spec.timing == (305, 335)
    assert spec.magnitude == (">", 183.0, 1)
    assert spec.rate_of_change == (">", 2.0, 1, 1, 0.0)
    assert spec.duration == (">", 3)


# ---- resolve_lake_csv_path ----------------------------------------------------------

@pytest.mark.parametrize("lake,expected_name", [
    ("superior", "superior_avg.csv"),
    ("michiganhuron", "michiganhuron_avg.csv"),
    ("stclair", "stclair_avg.csv"),
    ("erie", "erie_avg.csv"),
    ("ontario", "ontario_avg.csv"),
])
def test_resolve_lake_csv_path(lake, expected_name):
    data_dir = Path("some/data/dir")
    result = avg_batch_run.resolve_lake_csv_path(lake, data_dir)
    assert result == data_dir / expected_name


# ---- build_toml_text -----------------------------------------------------------------

def default_config(**overrides):
    config = avg_batch_run.BatchConfig()
    return replace(config, **overrides) if overrides else config


def parsed_toml(spec, timeseries_path=Path("data/clean/superior_avg.csv"),
                output_directory=Path("out/duluth_harbor/high_water"), config=None):
    text = avg_batch_run.build_toml_text(
        spec, timeseries_path, output_directory, config or default_config()
    )
    return text, tomllib.loads(text)


def test_build_toml_text_is_valid_toml_with_expected_sections():
    spec = avg_batch_run.parse_resource_row(base_row())
    _, data = parsed_toml(spec)
    assert "timeseries" in data
    assert "components" in data
    assert "output" in data


def test_build_toml_text_timeseries_section():
    spec = avg_batch_run.parse_resource_row(base_row())
    _, data = parsed_toml(
        spec, timeseries_path=Path("data/clean/superior_avg.csv"),
        config=default_config(first_day_of_water_year=91),
    )
    assert data["timeseries"]["path"] == "data/clean/superior_avg.csv"
    assert data["timeseries"]["first_day_of_water_year"] == 91


def test_build_toml_text_component_name_and_magnitude():
    spec = avg_batch_run.parse_resource_row(base_row())
    _, data = parsed_toml(spec)
    component = data["components"]["duluth_harbor_high_water"]
    assert component["magnitude"] == [">", 183.0, 1]


def test_build_toml_text_characteristic_key_order_fixed():
    row = base_row(
        timing_first_month=11, timing_last_month=12,
        rate_of_change_operator=">", rate_of_change_value=2.0,
        duration_operator=">", duration_value=3,
    )
    spec = avg_batch_run.parse_resource_row(row)
    _, data = parsed_toml(spec)
    component = data["components"]["duluth_harbor_high_water"]
    characteristic_keys = [
        k for k in component
        if k in ("timing", "magnitude", "rate_of_change", "duration")
    ]
    assert characteristic_keys == ["timing", "magnitude", "rate_of_change", "duration"]


def test_build_toml_text_omits_unused_characteristics():
    spec = avg_batch_run.parse_resource_row(base_row())  # only magnitude
    _, data = parsed_toml(spec)
    component = data["components"]["duluth_harbor_high_water"]
    assert "timing" not in component
    assert "rate_of_change" not in component
    assert "duration" not in component


def test_build_toml_text_verbose_written_first_when_false():
    row = base_row(verbose=False)
    spec = avg_batch_run.parse_resource_row(row)
    _, data = parsed_toml(spec)
    component = data["components"]["duluth_harbor_high_water"]
    assert component["verbose"] is False
    assert list(component.keys())[0] == "verbose"


def test_build_toml_text_success_pattern_always_emitted_even_when_true():
    row = base_row(success_pattern=True)
    spec = avg_batch_run.parse_resource_row(row)
    _, data = parsed_toml(spec)
    component = data["components"]["duluth_harbor_high_water"]
    assert component["success_pattern"] is True


def test_build_toml_text_success_pattern_written_last():
    row = base_row(success_pattern=False)
    spec = avg_batch_run.parse_resource_row(row)
    _, data = parsed_toml(spec)
    component = data["components"]["duluth_harbor_high_water"]
    assert component["success_pattern"] is False
    assert list(component.keys())[-1] == "success_pattern"


def test_build_toml_text_output_section():
    spec = avg_batch_run.parse_resource_row(base_row())
    output_dir = Path("out/duluth_harbor/high_water")
    _, data = parsed_toml(
        spec, output_directory=output_dir,
        config=default_config(overwrite=True, excel=False, metric_mode="return_period"),
    )
    assert data["output"]["directory"] == output_dir.as_posix()
    assert data["output"]["overwrite"] is True
    assert data["output"]["excel"] is False
    assert data["output"]["metric"]["mode"] == "return_period"


def test_build_toml_text_plot_enabled_by_default():
    spec = avg_batch_run.parse_resource_row(base_row())
    _, data = parsed_toml(spec)
    assert data["output"]["plot"]["enabled"] is True


def test_build_toml_text_plot_can_be_disabled_via_config():
    spec = avg_batch_run.parse_resource_row(base_row())
    config = default_config(plot_enabled=False)
    _, data = parsed_toml(spec, config=config)
    assert data["output"]["plot"]["enabled"] is False


def test_build_toml_text_plot_show_always_false_even_if_config_says_otherwise():
    spec = avg_batch_run.parse_resource_row(base_row())
    config = default_config(plot_enabled=True)
    _, data = parsed_toml(spec, config=config)
    assert data["output"]["plot"]["enabled"] is True
    assert data["output"]["plot"]["climate-canvas"]["show"] is False


def test_build_toml_text_plot_threshold_from_row_not_config():
    row = base_row(threshold=183.5)
    spec = avg_batch_run.parse_resource_row(row)
    config = default_config(plot_enabled=True)
    _, data = parsed_toml(spec, config=config)
    assert data["output"]["plot"]["climate-canvas"]["threshold"] == 183.5


def test_build_toml_text_no_threshold_key_when_row_threshold_blank():
    spec = avg_batch_run.parse_resource_row(base_row())
    config = default_config(plot_enabled=True)
    _, data = parsed_toml(spec, config=config)
    assert "threshold" not in data["output"]["plot"]["climate-canvas"]


def test_build_toml_text_plot_color_map_and_ticks():
    spec = avg_batch_run.parse_resource_row(base_row())
    config = default_config(
        plot_enabled=True, plot_color_map="viridis", plot_color_map_ticks=(-1.0, 0.0, 1.0)
    )
    _, data = parsed_toml(spec, config=config)
    climate_canvas = data["output"]["plot"]["climate-canvas"]
    assert climate_canvas["color_map"] == "viridis"
    assert climate_canvas["color_map_ticks"] == [-1.0, 0.0, 1.0]


def test_build_toml_text_no_color_map_ticks_key_when_none():
    spec = avg_batch_run.parse_resource_row(base_row())
    config = default_config(plot_enabled=True, plot_color_map_ticks=None)
    _, data = parsed_toml(spec, config=config)
    assert "color_map_ticks" not in data["output"]["plot"]["climate-canvas"]


# ---- resolve_output_folder ------------------------------------------------------------

def test_resolve_output_folder_flat_structure_is_base_directory():
    spec = avg_batch_run.parse_resource_row(base_row())
    config = default_config(output_directory="out", subdirectory_structure="flat")
    assert avg_batch_run.resolve_output_folder(spec, config) == Path("out")


def test_resolve_output_folder_resource_structure_nests_by_resource_name():
    spec = avg_batch_run.parse_resource_row(base_row())
    config = default_config(output_directory="out", subdirectory_structure="resource")
    assert avg_batch_run.resolve_output_folder(spec, config) == Path("out/duluth_harbor")


def test_resolve_output_folder_row_structure_nests_by_resource_and_component():
    spec = avg_batch_run.parse_resource_row(base_row())
    config = default_config(output_directory="out", subdirectory_structure="row")
    assert (
        avg_batch_run.resolve_output_folder(spec, config)
        == Path("out/duluth_harbor/high_water")
    )


def test_resolve_output_folder_invalid_subdirectory_structure_raises():
    spec = avg_batch_run.parse_resource_row(base_row())
    config = default_config(output_directory="out", subdirectory_structure="bogus")
    with pytest.raises(ValueError):
        avg_batch_run.resolve_output_folder(spec, config)


def test_resolve_output_folder_is_stateless_same_inputs_same_result():
    """No cross-row memory: two rows with identical resource/component resolve to the
    same folder (collision detection across rows is run_batch's job, not this
    function's -- see Step 5)."""
    spec_a = avg_batch_run.parse_resource_row(base_row())
    spec_b = avg_batch_run.parse_resource_row(base_row())
    config = default_config(output_directory="out", subdirectory_structure="row")
    assert (
        avg_batch_run.resolve_output_folder(spec_a, config)
        == avg_batch_run.resolve_output_folder(spec_b, config)
    )


# ---- read_resources_sheet / read_config_sheet -----------------------------------------

# pylint: disable-next=too-many-arguments,too-many-positional-arguments
def make_workbook(tmp_path, resources_header=None, resources_rows=None,
                  config_pairs=None, resources_sheet_name="resources",
                  config_sheet_name="config"):
    """Build a small .xlsx with a resources sheet and/or a config sheet for testing.

    resources_header: list of column names (row 1).
    resources_rows: list of tuples, one per data row (row 2+).
    config_pairs: list of (option, value) tuples (no header row -- see read_config_sheet,
                  which expects an 'option'/'value' header row, so callers pass that as
                  the first pair explicitly for full control, or rely on the default below).
    """
    from openpyxl import Workbook

    wb = Workbook()
    default_sheet = wb.active

    if resources_header is not None:
        ws = wb.create_sheet(resources_sheet_name) if default_sheet.title != resources_sheet_name \
            else default_sheet
        ws.title = resources_sheet_name
        ws.append(list(resources_header))
        for row in resources_rows or []:
            ws.append(list(row))
        default_sheet = None

    if config_pairs is not None:
        ws = wb.create_sheet(config_sheet_name) if default_sheet is None \
            else default_sheet
        ws.title = config_sheet_name
        ws.append(["option", "value"])
        for key, value in config_pairs:
            ws.append([key, value])
        default_sheet = None

    if default_sheet is not None and default_sheet.title == "Sheet":
        wb.remove(default_sheet)

    path = tmp_path / "workbook.xlsx"
    wb.save(path)
    return path


def test_read_resources_sheet_returns_one_dict_per_row(tmp_path):
    path = make_workbook(
        tmp_path,
        resources_header=["resource_name", "component_name", "lake", "magnitude_operator",
                          "magnitude_value"],
        resources_rows=[
            ("duluth_harbor", "high_water", "superior", ">", 183.0),
            ("green_bay", "low_water", "michiganhuron", "<", 175.0),
        ],
    )
    rows = avg_batch_run.read_resources_sheet(path)
    assert len(rows) == 2
    assert rows[0]["resource_name"] == "duluth_harbor"
    assert rows[1]["lake"] == "michiganhuron"


def test_read_resources_sheet_blank_cell_becomes_none(tmp_path):
    path = make_workbook(
        tmp_path,
        resources_header=["resource_name", "component_name", "lake", "threshold"],
        resources_rows=[("duluth_harbor", "high_water", "superior", None)],
    )
    rows = avg_batch_run.read_resources_sheet(path)
    assert rows[0]["threshold"] is None


def test_read_resources_sheet_missing_required_column_raises(tmp_path):
    path = make_workbook(
        tmp_path,
        resources_header=["resource_name", "lake"],  # missing component_name
        resources_rows=[("duluth_harbor", "superior")],
    )
    with pytest.raises(avg_batch_run.SheetValidationError) as exc_info:
        avg_batch_run.read_resources_sheet(path)
    assert "component_name" in str(exc_info.value)


def test_read_resources_sheet_unknown_column_raises(tmp_path):
    path = make_workbook(
        tmp_path,
        resources_header=["resource_name", "component_name", "lake", "not_a_real_column"],
        resources_rows=[("duluth_harbor", "high_water", "superior", "x")],
    )
    with pytest.raises(avg_batch_run.SheetValidationError) as exc_info:
        avg_batch_run.read_resources_sheet(path)
    assert "not_a_real_column" in str(exc_info.value)


def test_read_config_sheet_required_output_directory_only(tmp_path):
    path = make_workbook(tmp_path, config_pairs=[("output_directory", "out")])
    config = avg_batch_run.read_config_sheet(path)
    assert config.output_directory == "out"
    assert config == replace(avg_batch_run.BatchConfig(), output_directory="out")


def test_read_config_sheet_missing_output_directory_raises(tmp_path):
    path = make_workbook(tmp_path, config_pairs=[("metric_mode", "percentage")])
    with pytest.raises(avg_batch_run.SheetValidationError) as exc_info:
        avg_batch_run.read_config_sheet(path)
    assert "output_directory" in str(exc_info.value)


def test_read_config_sheet_unknown_option_raises(tmp_path):
    path = make_workbook(
        tmp_path, config_pairs=[("output_directory", "out"), ("not_a_real_option", "x")],
    )
    with pytest.raises(avg_batch_run.SheetValidationError) as exc_info:
        avg_batch_run.read_config_sheet(path)
    assert "not_a_real_option" in str(exc_info.value)


def test_read_config_sheet_duplicate_option_raises(tmp_path):
    path = make_workbook(
        tmp_path,
        config_pairs=[("output_directory", "out"), ("output_directory", "out2")],
    )
    with pytest.raises(avg_batch_run.SheetValidationError) as exc_info:
        avg_batch_run.read_config_sheet(path)
    assert "duplicate" in str(exc_info.value).lower()


@pytest.mark.parametrize("bad_value", ["bogus"])
def test_read_config_sheet_invalid_subdirectory_structure_raises(tmp_path, bad_value):
    path = make_workbook(
        tmp_path,
        config_pairs=[("output_directory", "out"), ("subdirectory_structure", bad_value)],
    )
    with pytest.raises(avg_batch_run.SheetValidationError) as exc_info:
        avg_batch_run.read_config_sheet(path)
    assert "subdirectory_structure" in str(exc_info.value)


def test_read_config_sheet_invalid_metric_mode_raises(tmp_path):
    path = make_workbook(
        tmp_path,
        config_pairs=[("output_directory", "out"), ("metric_mode", "bogus")],
    )
    with pytest.raises(avg_batch_run.SheetValidationError) as exc_info:
        avg_batch_run.read_config_sheet(path)
    assert "metric_mode" in str(exc_info.value)


def test_read_config_sheet_invalid_color_map_ticks_raises(tmp_path):
    path = make_workbook(
        tmp_path,
        config_pairs=[("output_directory", "out"), ("plot_color_map_ticks", "not,numbers")],
    )
    with pytest.raises(avg_batch_run.SheetValidationError) as exc_info:
        avg_batch_run.read_config_sheet(path)
    assert "plot_color_map_ticks" in str(exc_info.value)


def test_read_config_sheet_all_options_overridden(tmp_path):
    path = make_workbook(
        tmp_path,
        config_pairs=[
            ("output_directory", "out"),
            ("first_day_of_water_year", 91),
            ("metric_mode", "percentage"),
            ("excel", False),
            ("overwrite", True),
            ("subdirectory_structure", "row"),
            ("plot_enabled", False),
            ("plot_interpolate", False),
            ("plot_color_map", "viridis"),
            ("plot_color_map_ticks", "-1.0, 0.0, 1.0"),
        ],
    )
    config = avg_batch_run.read_config_sheet(path)
    assert config.output_directory == "out"
    assert config.first_day_of_water_year == 91
    assert config.metric_mode == "percentage"
    assert config.excel is False
    assert config.overwrite is True
    assert config.subdirectory_structure == "row"
    assert config.plot_enabled is False
    assert config.plot_interpolate is False
    assert config.plot_color_map == "viridis"
    assert config.plot_color_map_ticks == (-1.0, 0.0, 1.0)


def test_read_config_sheet_blank_optional_rows_use_defaults(tmp_path):
    path = make_workbook(
        tmp_path,
        config_pairs=[("output_directory", "out"), ("metric_mode", None)],
    )
    config = avg_batch_run.read_config_sheet(path)
    assert config.metric_mode == avg_batch_run.BatchConfig().metric_mode


# ---- run_batch --------------------------------------------------------------------------

def make_resources_workbook(tmp_path, rows, header=None):
    header = header or [
        "resource_name", "component_name", "lake", "magnitude_operator", "magnitude_value",
    ]
    return make_workbook(tmp_path, resources_header=header, resources_rows=rows)


def batch_config(tmp_path, **overrides):
    return replace(avg_batch_run.BatchConfig(), output_directory=str(tmp_path / "out"),
                   **overrides)


def test_run_batch_succeeds_and_writes_toml(tmp_path):
    resources_path = make_resources_workbook(
        tmp_path, [("duluth_harbor", "high_water", "superior", ">", 183.0)],
    )
    config = batch_config(tmp_path)
    calls = []

    summary = avg_batch_run.run_batch(
        resources_path, Path("data/clean"), config, run_toml=calls.append,
    )

    assert len(summary.succeeded) == 1
    assert len(summary.failed) == 0
    result = summary.results[0]
    assert result.row_index == 1
    assert result.resource_name == "duluth_harbor"
    assert result.status == "succeeded"
    expected_toml = Path(tmp_path / "out" / "duluth_harbor_high_water.toml")
    assert calls == [expected_toml]
    assert expected_toml.exists()


def test_run_batch_invalid_row_reported_as_failed_and_does_not_run(tmp_path):
    resources_path = make_resources_workbook(
        tmp_path, [("duluth_harbor", "", "superior", ">", 183.0)],  # missing component_name
    )
    config = batch_config(tmp_path)
    calls = []

    summary = avg_batch_run.run_batch(
        resources_path, Path("data/clean"), config, run_toml=calls.append,
    )

    assert len(summary.failed) == 1
    assert summary.failed[0].status == "failed"
    assert "component_name" in summary.failed[0].message
    assert not calls


def test_run_batch_continues_after_a_failed_row(tmp_path):
    resources_path = make_resources_workbook(
        tmp_path, [
            ("duluth_harbor", "", "superior", ">", 183.0),  # invalid: missing component_name
            ("green_bay", "low_water", "michiganhuron", "<", 175.0),  # valid
        ],
    )
    config = batch_config(tmp_path)
    calls = []

    summary = avg_batch_run.run_batch(
        resources_path, Path("data/clean"), config, run_toml=calls.append,
    )

    assert len(summary.results) == 2
    assert summary.results[0].status == "failed"
    assert summary.results[1].status == "succeeded"
    assert len(calls) == 1


def test_run_batch_duplicate_output_folder_reported_as_failed(tmp_path):
    resources_path = make_resources_workbook(
        tmp_path, [
            ("duluth_harbor", "high_water", "superior", ">", 183.0),
            ("duluth_harbor", "high_water", "superior", ">", 190.0),  # same folder ("row" mode)
        ],
    )
    config = batch_config(tmp_path, subdirectory_structure="row")
    calls = []

    summary = avg_batch_run.run_batch(
        resources_path, Path("data/clean"), config, run_toml=calls.append,
    )

    assert summary.results[0].status == "succeeded"
    assert summary.results[1].status == "failed"
    assert "Duplicate output folder" in summary.results[1].message
    assert len(calls) == 1


def test_run_batch_shared_flat_folder_with_distinct_components_both_succeed(tmp_path):
    """Flat mode (the default) intentionally shares one output folder across many
    rows -- only a real filename collision (same folder + same component_name) is a
    duplicate, not merely sharing a folder."""
    resources_path = make_resources_workbook(
        tmp_path, [
            ("duluth_harbor", "high_water", "superior", ">", 183.0),
            ("green_bay", "low_water", "michiganhuron", "<", 175.0),
        ],
    )
    config = batch_config(tmp_path)  # default subdirectory_structure="flat"
    calls = []

    summary = avg_batch_run.run_batch(
        resources_path, Path("data/clean"), config, run_toml=calls.append,
    )

    assert summary.results[0].status == "succeeded"
    assert summary.results[1].status == "succeeded"
    assert len(calls) == 2


def test_run_batch_pre_existing_output_fails_when_overwrite_false(tmp_path):
    resources_path = make_resources_workbook(
        tmp_path, [("duluth_harbor", "high_water", "superior", ">", 183.0)],
    )
    config = batch_config(tmp_path, overwrite=False)
    output_folder = Path(tmp_path / "out")
    output_folder.mkdir(parents=True)
    # pre-existing from a prior run
    (output_folder / "duluth_harbor_high_water_summary.xlsx").touch()
    calls = []

    summary = avg_batch_run.run_batch(
        resources_path, Path("data/clean"), config, run_toml=calls.append,
    )

    assert summary.results[0].status == "failed"
    assert "overwrite=False" in summary.results[0].message
    assert not calls


def test_run_batch_pre_existing_output_ignored_when_overwrite_true(tmp_path):
    resources_path = make_resources_workbook(
        tmp_path, [("duluth_harbor", "high_water", "superior", ">", 183.0)],
    )
    config = batch_config(tmp_path, overwrite=True)
    output_folder = Path(tmp_path / "out")
    output_folder.mkdir(parents=True)
    (output_folder / "high_water_summary.xlsx").touch()
    calls = []

    summary = avg_batch_run.run_batch(
        resources_path, Path("data/clean"), config, run_toml=calls.append,
    )

    assert summary.results[0].status == "succeeded"
    assert len(calls) == 1


def test_run_batch_unrelated_existing_files_do_not_block_a_row(tmp_path):
    """Only this row's own deterministic filenames matter, not sibling files (flat mode
    shares one output folder across many rows)."""
    resources_path = make_resources_workbook(
        tmp_path, [("duluth_harbor", "high_water", "superior", ">", 183.0)],
    )
    config = batch_config(tmp_path, overwrite=False)
    output_folder = Path(tmp_path / "out")
    output_folder.mkdir(parents=True)
    (output_folder / "some_other_component_summary.xlsx").touch()
    calls = []

    summary = avg_batch_run.run_batch(
        resources_path, Path("data/clean"), config, run_toml=calls.append,
    )

    assert summary.results[0].status == "succeeded"
    assert len(calls) == 1


def test_run_batch_run_toml_exception_reported_as_failed_row(tmp_path):
    resources_path = make_resources_workbook(
        tmp_path, [("duluth_harbor", "high_water", "superior", ">", 183.0)],
    )
    config = batch_config(tmp_path)

    def failing_run_toml(_path):
        raise RuntimeError("boom")

    summary = avg_batch_run.run_batch(
        resources_path, Path("data/clean"), config, run_toml=failing_run_toml,
    )

    assert summary.results[0].status == "failed"
    assert "boom" in summary.results[0].message


# ---- run_batch progress feedback ----------------------------------------------------

def test_run_batch_reports_progress_for_start_and_finish_of_each_row(tmp_path):
    resources_path = make_resources_workbook(
        tmp_path, [
            ("duluth_harbor", "high_water", "superior", ">", 183.0),
            ("bad_row", "", "superior", ">", 183.0),  # invalid: missing component_name
        ],
    )
    config = batch_config(tmp_path)
    progress_calls = []

    avg_batch_run.run_batch(
        resources_path, Path("data/clean"), config, run_toml=lambda _p: None,
        progress=lambda *args: progress_calls.append(args),
    )

    assert progress_calls == [
        (1, 2, "duluth_harbor", "high_water", "running"),
        (1, 2, "duluth_harbor", "high_water", "succeeded"),
        (2, 2, "bad_row", None, "running"),
        (2, 2, "bad_row", None, "failed"),
    ]


def test_default_progress_prints_row_index_total_and_status(capsys):
    avg_batch_run._default_progress(1, 3, "duluth_harbor", "high_water", "running")
    avg_batch_run._default_progress(1, 3, "duluth_harbor", "high_water", "succeeded")

    out = capsys.readouterr().out
    assert "[1/3] duluth_harbor/high_water: running" in out
    assert "[1/3] duluth_harbor/high_water: succeeded" in out


def test_expected_output_filenames_includes_optional_outputs_when_enabled():
    spec = avg_batch_run.parse_resource_row(base_row())
    config = default_config(excel=True, plot_enabled=True)
    filenames = avg_batch_run._expected_output_filenames(spec, config)
    assert filenames == [
        "duluth_harbor_high_water.toml", "duluth_harbor_high_water_summary.xlsx",
        "duluth_harbor_high_water_output.xlsx", "duluth_harbor_high_water_grid.csv",
        "duluth_harbor_high_water_plot.png",
    ]


def test_expected_output_filenames_excludes_optional_outputs_when_disabled():
    spec = avg_batch_run.parse_resource_row(base_row())
    config = default_config(excel=False, plot_enabled=False)
    filenames = avg_batch_run._expected_output_filenames(spec, config)
    assert filenames == ["duluth_harbor_high_water.toml", "duluth_harbor_high_water_summary.xlsx"]


def test_format_summary_reports_counts_and_failed_rows():
    summary = avg_batch_run.BatchSummary(results=[
        avg_batch_run.RowResult(1, "duluth_harbor", "high_water", "succeeded"),
        avg_batch_run.RowResult(2, "green_bay", "low_water", "failed", "boom"),
    ])
    text = avg_batch_run.format_summary(summary)
    assert "1 succeeded, 1 failed out of 2 row(s)." in text
    assert "Row 2" in text
    assert "boom" in text


# ---- run_batch integration test (real data, real hydropattern run) -----------------

REAL_DATA_DIR = Path(__file__).parent.parent / "examples" / "great_lakes" / "data" / "clean"


def test_run_batch_end_to_end_against_real_avg_csvs(tmp_path):
    """Runs a small fixture workbook through run_batch with the real hydropattern
    `run` (no mocking) against the real Great Lakes avg-level CSVs, verifying the
    whole pipeline (parse -> build .toml -> resolve output folder -> hydropattern
    run) actually produces the expected files on disk. Includes one invalid row to
    confirm the batch continues past it.
    """
    resources_path = make_resources_workbook(
        tmp_path,
        [
            ("duluth_harbor", "high_water", "superior", ">", 183.0),
            ("buffalo_shore", "low_water", "erie", "<", 174.0),
            ("bad_row", "", "erie", "<", 174.0),  # missing component_name -> invalid
        ],
    )
    config = batch_config(tmp_path, excel=False, plot_enabled=False)

    summary = avg_batch_run.run_batch(resources_path, REAL_DATA_DIR, config)

    assert len(summary.succeeded) == 2
    assert len(summary.failed) == 1
    assert summary.failed[0].resource_name == "bad_row"

    out_dir = tmp_path / "out"
    for qualified_name in ("duluth_harbor_high_water", "buffalo_shore_low_water"):
        assert (out_dir / f"{qualified_name}.toml").exists()
        assert (out_dir / f"{qualified_name}_summary.xlsx").exists()


# ---- main() CLI entry point --------------------------------------------------------

def test_main_returns_zero_and_prints_summary_when_all_rows_succeed(tmp_path, capsys, monkeypatch):
    resources_path = make_workbook(
        tmp_path,
        resources_header=[
            "resource_name", "component_name", "lake", "magnitude_operator", "magnitude_value",
        ],
        resources_rows=[("duluth_harbor", "high_water", "superior", ">", 183.0)],
        config_pairs=[("output_directory", str(tmp_path / "out"))],
    )
    monkeypatch.setattr(
        avg_batch_run, "run_batch",
        lambda *a, **k: avg_batch_run.BatchSummary(results=[
            avg_batch_run.RowResult(1, "duluth_harbor", "high_water", "succeeded"),
        ]),
    )

    exit_code = avg_batch_run.main([str(resources_path), str(tmp_path)])

    assert exit_code == 0
    assert "1 succeeded, 0 failed out of 1 row(s)." in capsys.readouterr().out


def test_main_returns_one_when_any_row_fails(tmp_path, monkeypatch):
    resources_path = make_workbook(
        tmp_path,
        resources_header=[
            "resource_name", "component_name", "lake", "magnitude_operator", "magnitude_value",
        ],
        resources_rows=[("duluth_harbor", "", "superior", ">", 183.0)],
        config_pairs=[("output_directory", str(tmp_path / "out"))],
    )
    monkeypatch.setattr(
        avg_batch_run, "run_batch",
        lambda *a, **k: avg_batch_run.BatchSummary(results=[
            avg_batch_run.RowResult(1, "duluth_harbor", None, "failed", "boom"),
        ]),
    )

    exit_code = avg_batch_run.main([str(resources_path), str(tmp_path)])

    assert exit_code == 1
