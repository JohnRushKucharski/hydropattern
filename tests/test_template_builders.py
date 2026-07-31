"""Tests for the Great Lakes template-builder scripts.

templates/build_avg_template.py and templates/build_twl_template.py live outside the
hydropattern package (examples/great_lakes/templates/ has no __init__.py), so they're
loaded here via importlib by file path, same as the other example scripts. Each
builder's own `build_template(path)` already accepts a destination path, so no
production refactor was needed to make these testable.
"""
import importlib.util
import sys
from pathlib import Path

import openpyxl

TEMPLATES_DIR = Path(__file__).parent.parent / "examples" / "great_lakes" / "templates"


def _load(module_name: str, filename: str):
    spec = importlib.util.spec_from_file_location(module_name, TEMPLATES_DIR / filename)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


build_avg_template = _load("build_avg_template", "build_avg_template.py")
build_twl_template = _load("build_twl_template", "build_twl_template.py")


# ---- build_avg_template.py ----------------------------------------------------------

def test_build_avg_template_writes_resources_and_config_sheets(tmp_path):
    out_path = build_avg_template.build_template(tmp_path / "template_avg.xlsx")

    wb = openpyxl.load_workbook(out_path)
    assert wb.sheetnames == ["resources", "config"]

    resources_ws = wb["resources"]
    header = [cell.value for cell in resources_ws[1]]
    assert header == build_avg_template.RESOURCES_HEADER

    # One example row is written after the header.
    assert resources_ws.max_row == 2
    example_row = [cell.value for cell in resources_ws[2]]
    assert example_row[header.index("resource_name")] == "duluth_harbor"
    assert example_row[header.index("lake")] == "superior"


def test_build_avg_template_config_sheet_has_all_options_with_defaults(tmp_path):
    out_path = build_avg_template.build_template(tmp_path / "template_avg.xlsx")

    wb = openpyxl.load_workbook(out_path)
    config_ws = wb["config"]
    rows = [(row[0].value, row[1].value) for row in config_ws.iter_rows(min_row=2)]

    # openpyxl reads a written "" cell back as None, so normalize before comparing.
    expected = [
        (option, default if default != "" else None)
        for option, default, _ in build_avg_template.CONFIG_ROWS
    ]
    assert rows == expected

    # output_directory has no sensible default and must be flagged as REQUIRED.
    output_dir_row = config_ws.cell(
        row=[r[0] for r in rows].index("output_directory") + 2, column=1
    )
    assert "REQUIRED" in output_dir_row.comment.text


# ---- build_twl_template.py -----------------------------------------------------------

def test_build_twl_template_writes_resources_and_config_sheets(tmp_path):
    out_path = build_twl_template.build_template(tmp_path / "template_twl.xlsx")

    wb = openpyxl.load_workbook(out_path)
    assert wb.sheetnames == ["resources", "config"]

    resources_ws = wb["resources"]
    header = [cell.value for cell in resources_ws[1]]
    assert header == build_twl_template.RESOURCES_HEADER

    # Three example rows illustrate save_point_id, lat/lon, and success_pattern=True.
    assert resources_ws.max_row == 4
    resource_names = [
        resources_ws.cell(row=r, column=header.index("resource_name") + 1).value
        for r in (2, 3, 4)
    ]
    assert resource_names == ["duluth_harbor", "mackinac_strait", "kingston_shoal"]


def test_build_twl_template_equivalent_elevation_column_has_explanatory_comment(tmp_path):
    out_path = build_twl_template.build_template(tmp_path / "template_twl.xlsx")

    wb = openpyxl.load_workbook(out_path)
    resources_ws = wb["resources"]
    header = [cell.value for cell in resources_ws[1]]
    col = header.index("equivalent_elevation") + 1
    comment = resources_ws.cell(row=1, column=col).comment
    assert comment is not None
    assert "baseline_magnitude" in comment.text


def test_build_twl_template_config_sheet_matches_config_rows(tmp_path):
    out_path = build_twl_template.build_template(tmp_path / "template_twl.xlsx")

    wb = openpyxl.load_workbook(out_path)
    config_ws = wb["config"]
    rows = [(row[0].value, row[1].value) for row in config_ws.iter_rows(min_row=2)]
    expected = [
        (option, default if default != "" else None)
        for option, default, _ in build_twl_template.CONFIG_ROWS
    ]
    assert rows == expected
