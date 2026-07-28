"""Tests for the Great Lakes batch_run_twl.py example script.

The script lives outside the hydropattern package (examples/great_lakes/), so it is
loaded here via importlib by file path rather than a normal package import.
"""
import importlib.util
import sys
from pathlib import Path

import pytest

SCRIPT_PATH = Path(__file__).parent.parent / "examples" / "great_lakes" / "batch_run_twl.py"
_spec = importlib.util.spec_from_file_location("twl_batch_run", SCRIPT_PATH)
assert _spec is not None and _spec.loader is not None
twl_batch_run = importlib.util.module_from_spec(_spec)
sys.modules["twl_batch_run"] = twl_batch_run
_spec.loader.exec_module(twl_batch_run)


# ---- _is_blank ----------------------------------------------------------------------

@pytest.mark.parametrize("value", [None, float("nan"), "", "   "])
def test_is_blank_true_cases(value):
    assert twl_batch_run._is_blank(value) is True


@pytest.mark.parametrize("value", [0, 0.0, "0", "x", "  x  ", False])
def test_is_blank_false_cases(value):
    assert twl_batch_run._is_blank(value) is False


# ---- _to_bool -------------------------------------------------------------------------

@pytest.mark.parametrize("value,default,expected", [
    (None, False, False),
    (None, True, True),
    ("", True, True),
    (True, False, True),
    (False, True, False),
    (1, False, True),
    (0, True, False),
])
def test_to_bool(value, default, expected):
    assert twl_batch_run._to_bool(value, default) is expected


# ---- _require_str -----------------------------------------------------------------

def test_require_str_returns_stripped_value():
    errors = []
    value = twl_batch_run._require_str({"resource_name": "  duluth_harbor  "},
                                        "resource_name", errors)
    assert value == "duluth_harbor"
    assert errors == []


@pytest.mark.parametrize("row", [{}, {"resource_name": None}, {"resource_name": "   "}])
def test_require_str_missing_or_blank_records_error(row):
    errors = []
    value = twl_batch_run._require_str(row, "resource_name", errors)
    assert value == ""
    assert len(errors) == 1
    assert "resource_name" in errors[0]


# ---- parse_scenario_sheet_name ------------------------------------------------------

@pytest.mark.parametrize("sheet_name,expected", [
    ("baseline-_0_0", "_0_0"),
    ("nearterm-_5_1.5", "_5_1.5"),
    ("moderate_low-_10_5", "_10_5"),
    ("extreme_low-_20_5", "_20_5"),
    ("extreme_high-_0_7", "_0_7"),
])
def test_parse_scenario_sheet_name_valid(sheet_name, expected):
    assert twl_batch_run.parse_scenario_sheet_name(sheet_name) == expected


@pytest.mark.parametrize("sheet_name", [
    "no_dash_here",
    "",
    "-_0_0",  # blank label
    "baseline_0_0",  # missing dash separator entirely
])
def test_parse_scenario_sheet_name_invalid_returns_none(sheet_name):
    assert twl_batch_run.parse_scenario_sheet_name(sheet_name) is None


# ---- interpolate_ari ------------------------------------------------------------------

ARIS = [0.1, 0.2, 0.5, 1, 2, 5, 10, 20, 50, 100, 200, 500, 1000]
LEVELS = [182.55, 183.18, 183.70, 183.80, 183.90, 183.98, 184.01,
          184.05, 184.09, 184.11, 184.13, 184.16, 184.18]


def test_interpolate_ari_exact_hit_on_a_known_point():
    ari = twl_batch_run.interpolate_ari(LEVELS, ARIS, 183.98)
    assert ari == pytest.approx(5)


def test_interpolate_ari_between_two_points():
    # Halfway between ARI=1 (183.80) and ARI=2 (183.90) -> ARI should be halfway (1.5).
    ari = twl_batch_run.interpolate_ari(LEVELS, ARIS, 183.85)
    assert ari == pytest.approx(1.5)


def test_interpolate_ari_below_minimum_clamps_and_warns():
    with pytest.warns(UserWarning, match="below curve minimum"):
        ari = twl_batch_run.interpolate_ari(LEVELS, ARIS, 100.0)
    assert ari == pytest.approx(0.1)


def test_interpolate_ari_above_maximum_clamps_and_warns():
    with pytest.warns(UserWarning, match="above curve maximum"):
        ari = twl_batch_run.interpolate_ari(LEVELS, ARIS, 500.0)
    assert ari == pytest.approx(1000)


def test_interpolate_ari_requires_matching_lengths():
    with pytest.raises(ValueError):
        twl_batch_run.interpolate_ari([1.0, 2.0], [1.0], 1.5)


def test_interpolate_ari_requires_at_least_two_points():
    with pytest.raises(ValueError):
        twl_batch_run.interpolate_ari([1.0], [1.0], 1.5)


# ---- exceedance_probability -----------------------------------------------------------

@pytest.mark.parametrize("operator,expected", [
    (">", 0.18126924692201818),     # p_exceed = 1 - exp(-1/ARI) = 1 - exp(-1/5)
    (">=", 0.18126924692201818),
    ("<", 0.8187307530779818),      # complement
    ("<=", 0.8187307530779818),
])
def test_exceedance_probability_operator_direction(operator, expected):
    p = twl_batch_run.exceedance_probability(LEVELS, ARIS, 183.98, operator)
    assert p == pytest.approx(expected)


@pytest.mark.parametrize("operator", ["=", "!=", "gt", ""])
def test_exceedance_probability_rejects_equality_operators(operator):
    with pytest.raises(ValueError, match=operator or "operator"):
        twl_batch_run.exceedance_probability(LEVELS, ARIS, 183.98, operator)


# ---- compute_metric ---------------------------------------------------------------

def test_compute_metric_portion_success_pattern_true():
    value = twl_batch_run.compute_metric(0.2, success_pattern=True, mode="portion")
    assert value == pytest.approx(0.2)


def test_compute_metric_portion_success_pattern_false_is_complement():
    value = twl_batch_run.compute_metric(0.2, success_pattern=False, mode="portion")
    assert value == pytest.approx(0.8)


def test_compute_metric_percentage():
    value = twl_batch_run.compute_metric(0.2, success_pattern=True, mode="percentage")
    assert value == pytest.approx(20.0)


def test_compute_metric_return_period():
    value = twl_batch_run.compute_metric(0.2, success_pattern=True, mode="return_period")
    assert value == pytest.approx(5.0)


def test_compute_metric_return_period_undefined_when_portion_zero():
    import math
    value = twl_batch_run.compute_metric(0.0, success_pattern=True, mode="return_period")
    assert math.isnan(value)


def test_compute_metric_rejects_unknown_mode():
    with pytest.raises(ValueError):
        twl_batch_run.compute_metric(0.2, success_pattern=True, mode="bogus")


# ---- select_save_point --------------------------------------------------------------

import pandas as pd

SAVE_POINTS = pd.DataFrame({
    "ID": [1, 2, 3],
    "lat": [46.50169, 46.48618, 46.76017],
    "lon": [-84.37344, -84.63203, -84.96525],
    0.1: [182.55, 182.62, 182.36],
})


def test_select_save_point_by_id():
    row = twl_batch_run.select_save_point(SAVE_POINTS, save_point_id=2)
    assert row["ID"] == 2
    assert row["lat"] == pytest.approx(46.48618)


def test_select_save_point_by_nearest_lat_lon():
    # closest to save point 3's coordinates (slightly offset)
    row = twl_batch_run.select_save_point(SAVE_POINTS, lat=46.75, lon=-84.95)
    assert row["ID"] == 3


def test_select_save_point_prefers_id_when_all_three_given():
    # lat/lon here are nearest to point 1, but id=3 should win.
    row = twl_batch_run.select_save_point(
        SAVE_POINTS, save_point_id=3, lat=46.50169, lon=-84.37344
    )
    assert row["ID"] == 3


def test_select_save_point_requires_id_or_lat_lon():
    with pytest.raises(ValueError, match="save_point_id.*lat.*lon|lat.*lon.*save_point_id"):
        twl_batch_run.select_save_point(SAVE_POINTS)


def test_select_save_point_requires_both_lat_and_lon():
    with pytest.raises(ValueError):
        twl_batch_run.select_save_point(SAVE_POINTS, lat=46.5)


def test_select_save_point_unknown_id_raises():
    with pytest.raises(ValueError, match="999"):
        twl_batch_run.select_save_point(SAVE_POINTS, save_point_id=999)


# ---- parse_resource_row -------------------------------------------------------------

def base_row(**overrides):
    row = {
        "resource_name": "duluth_harbor",
        "lake": "superior",
        "magnitude_operator": ">",
        "magnitude_value": 183.8,
        "save_point_id": 1,
    }
    row.update(overrides)
    return row


def test_parse_resource_row_minimal_valid_row():
    spec = twl_batch_run.parse_resource_row(base_row())
    assert spec.resource_name == "duluth_harbor"
    assert spec.component_name == "twl"
    assert spec.lake == "superior"
    assert spec.magnitude_operator == ">"
    assert spec.magnitude_value == pytest.approx(183.8)
    assert spec.save_point_id == 1
    assert spec.lat is None
    assert spec.lon is None
    assert spec.success_pattern is False
    assert spec.threshold is None
    assert spec.qualified_name == "duluth_harbor_twl"


def test_parse_resource_row_component_name_explicit():
    row = base_row(component_name="high_water")
    spec = twl_batch_run.parse_resource_row(row)
    assert spec.component_name == "high_water"
    assert spec.qualified_name == "duluth_harbor_high_water"


@pytest.mark.parametrize("missing", ["resource_name", "lake"])
def test_parse_resource_row_missing_required_field_raises(missing):
    row = base_row()
    row[missing] = ""
    with pytest.raises(twl_batch_run.RowValidationError):
        twl_batch_run.parse_resource_row(row)


def test_parse_resource_row_unknown_lake_raises():
    row = base_row(lake="erie")  # not a valid twl lake code
    with pytest.raises(twl_batch_run.RowValidationError) as exc_info:
        twl_batch_run.parse_resource_row(row)
    assert "lake" in str(exc_info.value).lower()


def test_parse_resource_row_magnitude_requires_value():
    row = base_row()
    del row["magnitude_value"]
    with pytest.raises(twl_batch_run.RowValidationError):
        twl_batch_run.parse_resource_row(row)


def test_parse_resource_row_magnitude_requires_operator():
    row = base_row()
    row["magnitude_operator"] = ""
    with pytest.raises(twl_batch_run.RowValidationError):
        twl_batch_run.parse_resource_row(row)


@pytest.mark.parametrize("operator", ["=", "!=", "~"])
def test_parse_resource_row_magnitude_invalid_operator_raises(operator):
    row = base_row(magnitude_operator=operator)
    with pytest.raises(twl_batch_run.RowValidationError):
        twl_batch_run.parse_resource_row(row)


def test_parse_resource_row_save_point_by_lat_lon():
    row = base_row(lat=46.5, lon=-84.4)
    del row["save_point_id"]
    spec = twl_batch_run.parse_resource_row(row)
    assert spec.save_point_id is None
    assert spec.lat == pytest.approx(46.5)
    assert spec.lon == pytest.approx(-84.4)


def test_parse_resource_row_save_point_id_wins_when_all_given():
    row = base_row(save_point_id=1, lat=46.5, lon=-84.4)
    spec = twl_batch_run.parse_resource_row(row)
    assert spec.save_point_id == 1
    assert spec.lat == pytest.approx(46.5)
    assert spec.lon == pytest.approx(-84.4)


def test_parse_resource_row_requires_id_or_lat_lon():
    row = base_row()
    del row["save_point_id"]
    with pytest.raises(twl_batch_run.RowValidationError) as exc_info:
        twl_batch_run.parse_resource_row(row)
    assert "save_point_id" in str(exc_info.value)


def test_parse_resource_row_requires_both_lat_and_lon():
    row = base_row(lat=46.5)
    del row["save_point_id"]
    with pytest.raises(twl_batch_run.RowValidationError):
        twl_batch_run.parse_resource_row(row)


@pytest.mark.parametrize("raw,expected", [(True, True), (False, False), ("", False), (None, False)])
def test_parse_resource_row_success_pattern_blank_defaults_false(raw, expected):
    row = base_row(success_pattern=raw)
    spec = twl_batch_run.parse_resource_row(row)
    assert spec.success_pattern is expected


def test_parse_resource_row_threshold_optional():
    row = base_row(threshold=183.5)
    spec = twl_batch_run.parse_resource_row(row)
    assert spec.threshold == pytest.approx(183.5)


def test_parse_resource_row_threshold_blank_is_none():
    spec = twl_batch_run.parse_resource_row(base_row())
    assert spec.threshold is None


# ---- resolve_lake_twl_path ------------------------------------------------------------

@pytest.mark.parametrize("lake,expected_name", [
    ("superior", "superior_twl.xlsx"),
    ("michigan", "michigan_twl.xlsx"),
    ("huron", "huron_twl.xlsx"),
    ("ontario", "ontario_twl.xlsx"),
])
def test_resolve_lake_twl_path(lake, expected_name):
    from pathlib import Path
    data_dir = Path("some/data/dir")
    result = twl_batch_run.resolve_lake_twl_path(lake, data_dir)
    assert result == data_dir / expected_name


# ---- resolve_output_folder ----------------------------------------------------------

def default_config(**overrides):
    from dataclasses import replace
    config = twl_batch_run.BatchConfig()
    return replace(config, **overrides) if overrides else config


def test_resolve_output_folder_flat():
    spec = twl_batch_run.parse_resource_row(base_row())
    config = default_config(output_directory="out", subdirectory_structure="flat")
    assert twl_batch_run.resolve_output_folder(spec, config) == Path("out")


def test_resolve_output_folder_resource():
    spec = twl_batch_run.parse_resource_row(base_row())
    config = default_config(output_directory="out", subdirectory_structure="resource")
    assert twl_batch_run.resolve_output_folder(spec, config) == Path("out/duluth_harbor")


def test_resolve_output_folder_row():
    spec = twl_batch_run.parse_resource_row(base_row())
    config = default_config(output_directory="out", subdirectory_structure="row")
    assert twl_batch_run.resolve_output_folder(spec, config) == Path("out/duluth_harbor/twl")


# ---- read_resources_sheet / read_config_sheet -----------------------------------------

def make_workbook(tmp_path, resources_header=None, resources_rows=None,
                  config_pairs=None, resources_sheet_name="resources",
                  config_sheet_name="config"):
    """Build a small .xlsx with a resources sheet and/or a config sheet for testing."""
    from openpyxl import Workbook

    wb = Workbook()
    default_sheet = wb.active

    if resources_header is not None:
        ws = wb.create_sheet(resources_sheet_name) if default_sheet.title != resources_sheet_name \
            else default_sheet
        ws.title = resources_sheet_name
        ws.append(list(resources_header))
        for row in resources_rows or []:
            ws.append(list(row))
        default_sheet = None

    if config_pairs is not None:
        ws = wb.create_sheet(config_sheet_name) if default_sheet is None \
            else default_sheet
        ws.title = config_sheet_name
        ws.append(["option", "value"])
        for key, value in config_pairs:
            ws.append([key, value])
        default_sheet = None

    if default_sheet is not None and default_sheet.title == "Sheet":
        wb.remove(default_sheet)

    path = tmp_path / "workbook.xlsx"
    wb.save(path)
    return path


def test_read_resources_sheet_returns_one_dict_per_row(tmp_path):
    path = make_workbook(
        tmp_path,
        resources_header=["resource_name", "lake", "magnitude_operator", "magnitude_value",
                          "save_point_id"],
        resources_rows=[
            ("duluth_harbor", "superior", ">", 183.8, 1),
            ("green_bay", "michigan", "<", 175.0, 2),
        ],
    )
    rows = twl_batch_run.read_resources_sheet(path)
    assert len(rows) == 2
    assert rows[0]["resource_name"] == "duluth_harbor"
    assert rows[1]["lake"] == "michigan"


def test_read_resources_sheet_blank_cell_becomes_none(tmp_path):
    path = make_workbook(
        tmp_path,
        resources_header=["resource_name", "lake", "magnitude_operator", "magnitude_value",
                          "threshold"],
        resources_rows=[("duluth_harbor", "superior", ">", 183.8, None)],
    )
    rows = twl_batch_run.read_resources_sheet(path)
    assert rows[0]["threshold"] is None


def test_read_resources_sheet_missing_required_column_raises(tmp_path):
    path = make_workbook(
        tmp_path,
        resources_header=["resource_name", "lake"],  # missing magnitude_operator/value
        resources_rows=[("duluth_harbor", "superior")],
    )
    with pytest.raises(twl_batch_run.SheetValidationError) as exc_info:
        twl_batch_run.read_resources_sheet(path)
    assert "magnitude_operator" in str(exc_info.value)


def test_read_resources_sheet_unknown_column_raises(tmp_path):
    path = make_workbook(
        tmp_path,
        resources_header=["resource_name", "lake", "magnitude_operator", "magnitude_value",
                          "not_a_real_column"],
        resources_rows=[("duluth_harbor", "superior", ">", 183.8, "x")],
    )
    with pytest.raises(twl_batch_run.SheetValidationError) as exc_info:
        twl_batch_run.read_resources_sheet(path)
    assert "not_a_real_column" in str(exc_info.value)


def test_read_config_sheet_required_output_directory_only(tmp_path):
    from dataclasses import replace
    path = make_workbook(tmp_path, config_pairs=[("output_directory", "out")])
    config = twl_batch_run.read_config_sheet(path)
    assert config.output_directory == "out"
    assert config == replace(twl_batch_run.BatchConfig(), output_directory="out")


def test_read_config_sheet_missing_output_directory_raises(tmp_path):
    path = make_workbook(tmp_path, config_pairs=[("metric_mode", "percentage")])
    with pytest.raises(twl_batch_run.SheetValidationError) as exc_info:
        twl_batch_run.read_config_sheet(path)
    assert "output_directory" in str(exc_info.value)


def test_read_config_sheet_unknown_option_raises(tmp_path):
    path = make_workbook(
        tmp_path, config_pairs=[("output_directory", "out"), ("not_a_real_option", "x")],
    )
    with pytest.raises(twl_batch_run.SheetValidationError) as exc_info:
        twl_batch_run.read_config_sheet(path)
    assert "not_a_real_option" in str(exc_info.value)


def test_read_config_sheet_duplicate_option_raises(tmp_path):
    path = make_workbook(
        tmp_path,
        config_pairs=[("output_directory", "out"), ("output_directory", "out2")],
    )
    with pytest.raises(twl_batch_run.SheetValidationError) as exc_info:
        twl_batch_run.read_config_sheet(path)
    assert "duplicate" in str(exc_info.value).lower()


def test_read_config_sheet_invalid_subdirectory_structure_raises(tmp_path):
    path = make_workbook(
        tmp_path,
        config_pairs=[("output_directory", "out"), ("subdirectory_structure", "bogus")],
    )
    with pytest.raises(twl_batch_run.SheetValidationError) as exc_info:
        twl_batch_run.read_config_sheet(path)
    assert "subdirectory_structure" in str(exc_info.value)


def test_read_config_sheet_invalid_metric_mode_raises(tmp_path):
    path = make_workbook(
        tmp_path,
        config_pairs=[("output_directory", "out"), ("metric_mode", "bogus")],
    )
    with pytest.raises(twl_batch_run.SheetValidationError) as exc_info:
        twl_batch_run.read_config_sheet(path)
    assert "metric_mode" in str(exc_info.value)


def test_read_config_sheet_invalid_color_map_ticks_raises(tmp_path):
    path = make_workbook(
        tmp_path,
        config_pairs=[("output_directory", "out"), ("plot_color_map_ticks", "not,numbers")],
    )
    with pytest.raises(twl_batch_run.SheetValidationError) as exc_info:
        twl_batch_run.read_config_sheet(path)
    assert "plot_color_map_ticks" in str(exc_info.value)


def test_read_config_sheet_all_options_overridden(tmp_path):
    path = make_workbook(
        tmp_path,
        config_pairs=[
            ("output_directory", "out"),
            ("metric_mode", "percentage"),
            ("overwrite", True),
            ("subdirectory_structure", "row"),
            ("plot_interpolate", False),
            ("plot_color_map", "viridis"),
            ("plot_color_map_ticks", "-1.0, 0.0, 1.0"),
        ],
    )
    config = twl_batch_run.read_config_sheet(path)
    assert config.output_directory == "out"
    assert config.metric_mode == "percentage"
    assert config.overwrite is True
    assert config.subdirectory_structure == "row"
    assert config.plot_interpolate is False
    assert config.plot_color_map == "viridis"
    assert config.plot_color_map_ticks == (-1.0, 0.0, 1.0)


def test_read_config_sheet_blank_optional_rows_use_defaults(tmp_path):
    path = make_workbook(
        tmp_path,
        config_pairs=[("output_directory", "out"), ("metric_mode", None)],
    )
    config = twl_batch_run.read_config_sheet(path)
    assert config.metric_mode == twl_batch_run.BatchConfig().metric_mode


# ---- compute_scenario_metrics / build_resource_outputs ---------------------------------

def make_twl_workbook(path, sheet_frames: dict):
    """Write a small multi-sheet twl-like workbook (columns ID, lat, lon, <ARI...>)."""
    with pd.ExcelWriter(path) as writer:
        for name, df in sheet_frames.items():
            df.to_excel(writer, sheet_name=name, index=False)
    return path


def _twl_sheet_frame(save_point_id=1, lat=10.0, lon=20.0, levels=(100.0, 110.0, 115.0, 120.0)):
    return pd.DataFrame({
        "ID": [save_point_id],
        "lat": [lat],
        "lon": [lon],
        1: [levels[0]],
        5: [levels[1]],
        10: [levels[2]],
        50: [levels[3]],
    })


GRID_SHEET_NAMES = ["baseline-_0_0", "a-_0_1.5", "b-_5_0", "c-_5_1.5"]


def test_compute_scenario_metrics_one_value_per_grid_sheet(tmp_path):
    path = make_twl_workbook(
        tmp_path / "lake_twl.xlsx",
        {name: _twl_sheet_frame() for name in GRID_SHEET_NAMES},
    )
    resource = twl_batch_run.parse_resource_row(base_row(magnitude_operator=">", magnitude_value=110.0))
    values = twl_batch_run.compute_scenario_metrics(resource, path, "portion")
    assert set(values.keys()) == {"_0_0", "_0_1.5", "_5_0", "_5_1.5"}
    assert all(0.0 <= v <= 1.0 for v in values.values())


def test_compute_scenario_metrics_skips_non_grid_sheets(tmp_path):
    path = make_twl_workbook(
        tmp_path / "lake_twl.xlsx",
        {"not_a_grid_sheet": _twl_sheet_frame(), "baseline-_0_0": _twl_sheet_frame()},
    )
    resource = twl_batch_run.parse_resource_row(base_row())
    values = twl_batch_run.compute_scenario_metrics(resource, path, "portion")
    assert set(values.keys()) == {"_0_0"}


def _recording_plot_fn(calls):
    def plot_fn(xs, ys, zs, **kwargs):
        calls.append({"xs": xs, "ys": ys, "zs": zs, **kwargs})
    return plot_fn


def test_build_resource_outputs_writes_grid_csv_and_calls_plot_fn(tmp_path):
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    make_twl_workbook(
        data_dir / "superior_twl.xlsx",
        {name: _twl_sheet_frame() for name in GRID_SHEET_NAMES},
    )
    resource = twl_batch_run.parse_resource_row(base_row())
    config = default_config(output_directory=str(tmp_path / "out"))
    calls = []
    grid_path, plot_path = twl_batch_run.build_resource_outputs(
        resource, data_dir, config, plot_fn=_recording_plot_fn(calls)
    )
    assert grid_path == tmp_path / "out" / "duluth_harbor_twl_grid.csv"
    assert plot_path == tmp_path / "out" / "duluth_harbor_twl_plot.png"
    assert grid_path.exists()
    assert len(calls) == 1
    assert calls[0]["title"] == "duluth_harbor_twl"
    assert calls[0]["save_path"] == plot_path
    assert calls[0]["show"] is False


def test_build_resource_outputs_overwrite_false_raises_on_existing(tmp_path):
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    make_twl_workbook(
        data_dir / "superior_twl.xlsx",
        {name: _twl_sheet_frame() for name in GRID_SHEET_NAMES},
    )
    resource = twl_batch_run.parse_resource_row(base_row())
    config = default_config(output_directory=str(tmp_path / "out"), overwrite=False)
    twl_batch_run.build_resource_outputs(resource, data_dir, config, plot_fn=_recording_plot_fn([]))
    with pytest.raises(FileExistsError):
        twl_batch_run.build_resource_outputs(
            resource, data_dir, config, plot_fn=_recording_plot_fn([])
        )


def test_build_resource_outputs_overwrite_true_replaces(tmp_path):
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    make_twl_workbook(
        data_dir / "superior_twl.xlsx",
        {name: _twl_sheet_frame() for name in GRID_SHEET_NAMES},
    )
    resource = twl_batch_run.parse_resource_row(base_row())
    config = default_config(output_directory=str(tmp_path / "out"), overwrite=True)
    twl_batch_run.build_resource_outputs(resource, data_dir, config, plot_fn=_recording_plot_fn([]))
    # Should not raise the second time.
    twl_batch_run.build_resource_outputs(resource, data_dir, config, plot_fn=_recording_plot_fn([]))


# ---- run_batch --------------------------------------------------------------------------

def test_run_batch_all_rows_succeed(tmp_path):
    resources_path = make_workbook(
        tmp_path,
        resources_header=["resource_name", "lake", "magnitude_operator", "magnitude_value",
                          "save_point_id"],
        resources_rows=[
            ("duluth_harbor", "superior", ">", 110.0, 1),
            ("green_bay", "michigan", "<", 175.0, 27),
        ],
    )
    config = twl_batch_run.BatchConfig(output_directory=str(tmp_path / "out"))
    calls = []

    def fake_build_outputs(resource, data_dir, cfg):
        calls.append(resource.resource_name)
        return (Path("grid.csv"), Path("plot.png"))

    summary = twl_batch_run.run_batch(
        resources_path, tmp_path, config, build_outputs=fake_build_outputs,
        progress=lambda *a: None,
    )
    assert len(summary.succeeded) == 2
    assert len(summary.failed) == 0
    assert calls == ["duluth_harbor", "green_bay"]


def test_run_batch_continues_past_row_validation_failure(tmp_path):
    resources_path = make_workbook(
        tmp_path,
        resources_header=["resource_name", "lake", "magnitude_operator", "magnitude_value",
                          "save_point_id"],
        resources_rows=[
            ("", "superior", ">", 110.0, 1),  # invalid: blank resource_name
            ("green_bay", "michigan", "<", 175.0, 27),
        ],
    )
    config = twl_batch_run.BatchConfig(output_directory=str(tmp_path / "out"))
    summary = twl_batch_run.run_batch(
        resources_path, tmp_path, config,
        build_outputs=lambda r, d, c: (Path("g.csv"), Path("p.png")),
        progress=lambda *a: None,
    )
    assert len(summary.failed) == 1
    assert len(summary.succeeded) == 1
    assert summary.results[0].row_index == 1
    assert summary.results[0].status == "failed"


def test_run_batch_continues_past_build_outputs_error(tmp_path):
    resources_path = make_workbook(
        tmp_path,
        resources_header=["resource_name", "lake", "magnitude_operator", "magnitude_value",
                          "save_point_id"],
        resources_rows=[
            ("duluth_harbor", "superior", ">", 110.0, 1),
            ("green_bay", "michigan", "<", 175.0, 27),
        ],
    )
    config = twl_batch_run.BatchConfig(output_directory=str(tmp_path / "out"))

    def flaky_build_outputs(resource, data_dir, cfg):
        if resource.resource_name == "duluth_harbor":
            raise ValueError("boom")
        return (Path("g.csv"), Path("p.png"))

    summary = twl_batch_run.run_batch(
        resources_path, tmp_path, config, build_outputs=flaky_build_outputs,
        progress=lambda *a: None,
    )
    assert len(summary.failed) == 1
    assert summary.failed[0].message == "boom"
    assert len(summary.succeeded) == 1


def test_run_batch_duplicate_output_target_fails_second_row(tmp_path):
    resources_path = make_workbook(
        tmp_path,
        resources_header=["resource_name", "lake", "magnitude_operator", "magnitude_value",
                          "save_point_id"],
        resources_rows=[
            ("duluth_harbor", "superior", ">", 110.0, 1),
            ("duluth_harbor", "michigan", "<", 175.0, 27),  # same resource+component name
        ],
    )
    config = twl_batch_run.BatchConfig(
        output_directory=str(tmp_path / "out"), subdirectory_structure="flat",
    )
    summary = twl_batch_run.run_batch(
        resources_path, tmp_path, config,
        build_outputs=lambda r, d, c: (Path("g.csv"), Path("p.png")),
        progress=lambda *a: None,
    )
    assert len(summary.succeeded) == 1
    assert len(summary.failed) == 1
    assert "Duplicate output target" in summary.failed[0].message


def test_format_summary_includes_failure_details(tmp_path):
    resources_path = make_workbook(
        tmp_path,
        resources_header=["resource_name", "lake", "magnitude_operator", "magnitude_value",
                          "save_point_id"],
        resources_rows=[("", "superior", ">", 110.0, 1)],
    )
    config = twl_batch_run.BatchConfig(output_directory=str(tmp_path / "out"))
    summary = twl_batch_run.run_batch(
        resources_path, tmp_path, config,
        build_outputs=lambda r, d, c: (Path("g.csv"), Path("p.png")),
        progress=lambda *a: None,
    )
    text = twl_batch_run.format_summary(summary)
    assert "1 succeeded, 1 failed" not in text  # sanity: only 1 row total here
    assert "0 succeeded, 1 failed" in text
    assert "Row 1" in text
